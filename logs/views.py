"""
Logs App - Views
"""

from django.views.generic import ListView
from django.http import HttpResponse
from .models import ActivityLog
from accounts.mixins import CommanderOrGHeadMixin


class ActivityLogListView(CommanderOrGHeadMixin, ListView):
    model = ActivityLog
    template_name = 'logs/activity_log.html'
    context_object_name = 'logs'
    paginate_by = 50

    def get_queryset(self):
        # Exclude locked sheet actions (LOCK and UNLOCK) from display
        queryset = ActivityLog.objects.select_related('user').exclude(action__in=['LOCK', 'UNLOCK'])
        action = self.request.GET.get('action')
        user_id = self.request.GET.get('user')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if action:
            queryset = queryset.filter(action=action)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # Exclude LOCK and UNLOCK from choices
        ctx['action_choices'] = [
            choice for choice in ActivityLog.ACTION_CHOICES 
            if choice[0] not in ['LOCK', 'UNLOCK']
        ]

        # Calculate dynamic database counts for each action, matching search filters
        from django.db.models import Count
        user_id = self.request.GET.get('user')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        base_qs = ActivityLog.objects.exclude(action__in=['LOCK', 'UNLOCK'])
        if user_id:
            base_qs = base_qs.filter(user_id=user_id)
        if date_from:
            base_qs = base_qs.filter(timestamp__date__gte=date_from)
        if date_to:
            base_qs = base_qs.filter(timestamp__date__lte=date_to)
            
        ctx['total_count'] = base_qs.count()
        
        db_counts = dict(
            base_qs.values('action').annotate(c=Count('id')).values_list('action', 'c')
        )
        
        ctx['action_counts'] = {
            choice[0]: db_counts.get(choice[0], 0) 
            for choice in ctx['action_choices']
        }

        return ctx


class ExportLogsView(CommanderOrGHeadMixin, ListView):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Apply same filters as the list view, excluding LOCK and UNLOCK
        action = request.GET.get('action')
        user_id = request.GET.get('user')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

        queryset = ActivityLog.objects.select_related('user').exclude(action__in=['LOCK', 'UNLOCK'])
        if action:
            queryset = queryset.filter(action=action)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)

        # Retrieve up to 2000 records for Excel export
        logs = queryset[:2000]

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activity Logs"

        # Enable Gridlines
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')

        data_font = Font(name='Calibri', size=10, color='1A1A1A')
        
        # Borders
        thin_side = Side(style='thin', color='C3D9CE')
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Alignments
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Zebra striping fills
        row_fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        row_fill_odd = PatternFill(start_color='F4F9F6', end_color='F4F9F6', fill_type='solid')

        # Action styles mapping (font, fill)
        action_styles = {
            'LOGIN': {
                'font': Font(name='Calibri', size=10, bold=True, color='1B5E20'),
                'fill': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            },
            'LOGOUT': {
                'font': Font(name='Calibri', size=10, bold=True, color='B71C1C'),
                'fill': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            },
            'CREATE': {
                'font': Font(name='Calibri', size=10, bold=True, color='0D47A1'),
                'fill': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            },
            'UPDATE': {
                'font': Font(name='Calibri', size=10, bold=True, color='E65100'),
                'fill': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
            },
            'DELETE': {
                'font': Font(name='Calibri', size=10, bold=True, color='880E4F'),
                'fill': PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
            },
            'EXPORT': {
                'font': Font(name='Calibri', size=10, bold=True, color='4A148C'),
                'fill': PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
            },
            'EVALUATE': {
                'font': Font(name='Calibri', size=10, bold=True, color='004D40'),
                'fill': PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
            },
        }
        default_action_style = {
            'font': Font(name='Calibri', size=10, bold=True, color='37474F'),
            'fill': PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid')
        }

        # Write Title
        headers = ['S.No', 'Timestamp', 'User', 'Role', 'Action', 'Description', 'IP Address', 'User Agent']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="SYSTEM ACTIVITY LOGS")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 36

        # Write Column Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[2].height = 26

        # Write Data rows
        for row_idx, log in enumerate(logs, 3):
            # Select row fill for zebra striping
            r_fill = row_fill_odd if row_idx % 2 == 1 else row_fill_even
            
            s_no = row_idx - 2
            timestamp_str = log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            user_str = log.user.username if log.user else 'System'
            role_str = log.role or 'N/A'
            action_display = log.get_action_display()
            description_str = log.description
            ip_str = log.ip_address or 'N/A'
            
            # Simple User Agent OS parser
            ua_raw = (log.user_agent or '').lower()
            if 'windows' in ua_raw:
                ua_str = 'Windows'
            elif 'ipad' in ua_raw:
                ua_str = 'iPad (iOS)'
            elif 'iphone' in ua_raw:
                ua_str = 'iPhone (iOS)'
            elif 'macintosh' in ua_raw or 'mac os' in ua_raw:
                ua_str = 'macOS'
            elif 'android' in ua_raw:
                ua_str = 'Android'
            elif 'linux' in ua_raw:
                ua_str = 'Linux'
            elif log.user_agent:
                ua_str = 'Other'
            else:
                ua_str = 'N/A'

            row_data = [s_no, timestamp_str, user_str, role_str, action_display, description_str, ip_str, ua_str]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = r_fill
                cell.border = border

                # Alignments
                if col_idx in [1, 2, 3, 4, 7]: # S.No, Timestamp, User, Role, IP
                    cell.alignment = center_align
                else: # Action, Description, User Agent
                    cell.alignment = left_align

                # Special Action status pill styling
                if col_idx == 5: # Action Column
                    cell.alignment = center_align
                    style = action_styles.get(log.action, default_action_style)
                    cell.font = style['font']
                    cell.fill = style['fill']

            ws.row_dimensions[row_idx].height = 22

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                # Skip title cell in first row for width calculation
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            # Limit description and user agent columns from getting too wide
            if col[0].column == 6: # Description
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 50)
            elif col[0].column == 8: # User Agent
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

        # Build HTTP Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="activity_logs.xlsx"'
        wb.save(response)
        return response
