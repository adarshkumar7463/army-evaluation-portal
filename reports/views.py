"""
Reports App - Views
PDF, Excel, and CSV export functionality with department-specific access control
"""

import csv
import io
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, FileResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils import timezone
from django.db.models import Q, Sum, F, Value, FloatField
from django.db.models.functions import Coalesce

from departments.models import Agniveer
from evaluation.models import EvaluationSheet, Marks
from evaluation.result_helpers import (
    build_tts_result_row,
    build_battalion_result_row,
    build_department_result_row,
    build_cs_result_row,
    build_clerk_result_row,
    get_bn_desp_list,
    get_bn_desp_q,
    ALL_BATTALION_UNITS_LIST
)
from accounts.models import CustomUser
from accounts.mixins import AnyStaffMixin, CommanderOrGHeadMixin
from logs.utils import log_action


DEPARTMENT_NAMES = {'A': 'Battalion', 'B': 'TTS', 'C': 'CS', 'D': 'Clerk'}
CLERK_TRADES = ['CLK', 'CLERK', 'Clerk', 'CLK_SD', 'CLK_IM']
TTS_DIRECT_TRADES = ['DMV', 'OPEM']


def scoped_agniveers(queryset, user, dept=None):
    dept = dept or user.get_department_code()
    if dept == 'A':
        if user.is_department and user.battalion_unit:
            return queryset.filter(get_bn_desp_q('bn_desp', user.battalion_unit))
        battalion_units = get_bn_desp_list([choice[0] for choice in CustomUser.BATTALION_CHOICES])
        return queryset.filter(bn_desp__in=battalion_units)
    if dept == 'B':
        if user.is_department and user.tts_trade == 'DMV':
            return queryset.filter(trade='DMV')
        if user.is_department and user.tts_trade == 'OPEM':
            return queryset.filter(trade='OPEM')
        if user.is_department and user.tts_trade == 'OTHER':
            return queryset.exclude(trade__in=TTS_DIRECT_TRADES + CLERK_TRADES)
        return queryset.exclude(trade__in=CLERK_TRADES)
    if dept == 'C':
        return queryset
    if dept == 'D':
        return queryset.filter(trade__in=CLERK_TRADES)
    return queryset


def scoped_sheets(queryset, user, dept=None):
    dept = dept or user.get_department_code()
    queryset = queryset.filter(department=dept)
    if dept == 'A':
        if user.is_department and user.battalion_unit:
            return queryset.filter(get_bn_desp_q('agniveer__bn_desp', user.battalion_unit))
        battalion_units = get_bn_desp_list([choice[0] for choice in CustomUser.BATTALION_CHOICES])
        return queryset.filter(agniveer__bn_desp__in=battalion_units)
    if dept == 'B':
        if user.is_department and user.tts_trade == 'DMV':
            return queryset.filter(agniveer__trade='DMV')
        if user.is_department and user.tts_trade == 'OPEM':
            return queryset.filter(agniveer__trade='OPEM')
        if user.is_department and user.tts_trade == 'OTHER':
            return queryset.exclude(agniveer__trade__in=TTS_DIRECT_TRADES)
        return queryset
    if dept == 'D':
        return queryset.filter(agniveer__trade__in=CLERK_TRADES)
    return queryset


def user_can_access_agniveer(user, agniveer, dept=None):
    return scoped_agniveers(Agniveer.objects.filter(pk=agniveer.pk), user, dept).exists()


def pass_fail_counts_for_scope(user, dept=None):
    from evaluation.result_helpers import is_sheet_evaluated
    departments = [dept] if dept else ['A']
    all_sheets = EvaluationSheet.objects.all().prefetch_related('marks')

    if dept:
        all_sheets = scoped_sheets(all_sheets, user, dept)
        agniveers = scoped_agniveers(Agniveer.objects.all(), user, dept)
    else:
        agniveers = Agniveer.objects.all()

    passed = 0
    failed = 0
    for agniveer in agniveers:
        total_marks = 0
        max_marks = 0
        for dept_code in departments:
            dept_sheets = [s for s in all_sheets.filter(agniveer=agniveer, department=dept_code) if is_sheet_evaluated(s)]
            if not dept_sheets:
                continue
            result_row = build_department_result_row(agniveer, dept_sheets, dept_code)
            total_marks += result_row.get('grand_total', 0) or 0
            max_marks += result_row.get('max_total') or 40
        if max_marks <= 0:
            continue
        percentage = (total_marks / max_marks) * 100
        passing_threshold = 40 if 'A' in departments and len(departments) == 1 else 50
        if percentage >= passing_threshold:
            passed += 1
        else:
            failed += 1

    return {
        'evaluated': passed + failed,
        'passed': passed,
        'failed': failed,
    }

# Department-specific access mixins
class TTSTradeHeadMixin(UserPassesTestMixin):
    """Allow access only to TTS Trade Head users"""
    def test_func(self):
        return self.request.user.is_authenticated and getattr(self.request.user, 'tts_trade', None) is not None

class BattalionHeadMixin(UserPassesTestMixin):
    """Allow access only to Battalion Head users"""
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_battalion

class CSHeadMixin(UserPassesTestMixin):
    """Allow access only to CS head users"""
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role == CustomUser.ROLE_DEPT_C

class ClerkHeadMixin(UserPassesTestMixin):
    """Allow access only to Clerk head users"""
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role == CustomUser.ROLE_DEPT_D


class ReportDashboardView(AnyStaffMixin, View):
    template_name = 'reports/report_dashboard.html'

    def get(self, request):
        user = request.user
        context = {}
        
        if user.is_commander or user.is_g_head:
            # Commander/G-Head sees ALL departments
            context['user_role'] = 'commander_ghead'
            context['show_all_departments'] = True
            context['report_departments'] = ['A', 'B', 'C', 'D']
            
            # Get data for all departments
            dept_data = {}
            for dept_code, dept_name in DEPARTMENT_NAMES.items():
                counts = pass_fail_counts_for_scope(user, dept_code)
                dept_data[dept_code] = {
                    'name': dept_name,
                    'total_agniveers': scoped_agniveers(Agniveer.objects.all(), user, dept_code).count(),
                    'evaluated_agniveers': counts['evaluated'],
                    'pass_count': counts['passed'],
                    'fail_count': counts['failed'],
                    'total_trainers': CustomUser.objects.filter(
                        role__in=[CustomUser.ROLE_TRAINER_NCO, CustomUser.ROLE_TRAINER_JCO, CustomUser.ROLE_TRAINER_OFFICER],
                        department=dept_code
                    ).count()
                }
            context['department_data'] = dept_data
            
            # Overall totals
            counts = pass_fail_counts_for_scope(user)
            context['total_agniveers'] = Agniveer.objects.count()
            context['total_trainers'] = CustomUser.objects.filter(
                role__in=[CustomUser.ROLE_TRAINER_NCO, CustomUser.ROLE_TRAINER_JCO, CustomUser.ROLE_TRAINER_OFFICER]
            ).count()
            context['evaluated_agniveers'] = counts['evaluated']
            context['pass_count'] = counts['passed']
            context['fail_count'] = counts['failed']
            
        elif user.is_department:
            dept = user.get_department_code()
            context['user_role'] = f'department_{dept}'
            context['department_code'] = dept
            context['department_name'] = DEPARTMENT_NAMES.get(dept, dept)
            context['report_departments'] = [dept]

            context['is_department_head'] = True
            if dept == 'A':
                context['head_type'] = 'battalion'
            elif dept == 'B':
                context['head_type'] = 'tts'
                context['tts_trade'] = user.tts_trade
            elif dept == 'C':
                context['head_type'] = 'cs'
            elif dept == 'D':
                context['head_type'] = 'clerk'

            agniveers = scoped_agniveers(Agniveer.objects.all(), user, dept)
            counts = pass_fail_counts_for_scope(user, dept)
            
            context['total_agniveers'] = agniveers.count()
            context['total_trainers'] = CustomUser.objects.filter(
                role__in=[CustomUser.ROLE_TRAINER_NCO, CustomUser.ROLE_TRAINER_JCO, CustomUser.ROLE_TRAINER_OFFICER],
                department=dept
            ).count()
            context['evaluated_agniveers'] = counts['evaluated']
            context['pass_count'] = counts['passed']
            context['fail_count'] = counts['failed']
        else:
            # Non-department users (trainers hidden): provide empty/default report data
            context['user_role'] = 'trainer'
            context['report_departments'] = []
            context['total_agniveers'] = 0
            context['total_trainers'] = 0
            context['pass_count'] = 0
            context['evaluated_agniveers'] = 0

        context['fail_count'] = context.get('fail_count', 0)

        # Provide available test types to templates for per-test exports
        try:
            test_types = list(EvaluationSheet.TEST_TYPE_CHOICES)
        except Exception:
            test_types = []
        context['test_types'] = test_types

        return render(request, self.template_name, context)


# Department-specific report card views
class TTSReportCardView(TTSTradeHeadMixin, View):
    """TTS Trade Head sees report cards for their specific trade only"""
    template_name = 'reports/tts_report_card.html'
    
    def get(self, request):
        user = request.user
        tts_trade = user.tts_trade
        
        # Filter agniveers by trade
        if tts_trade == 'DMV':
            agniveers = Agniveer.objects.filter(trade='DMV')
        elif tts_trade == 'OPEM':
            agniveers = Agniveer.objects.filter(trade='OPEM')
        else:  # OTHER
            agniveers = Agniveer.objects.exclude(trade__in=['DMV', 'OPEM'])
        
        sheets = EvaluationSheet.objects.filter(
            department='B', 
            is_locked=True,
            agniveer__in=agniveers
        ).select_related('agniveer').prefetch_related('marks')
        
        # Build result rows using existing TTS result helper
        result_rows = []
        for agniveer in agniveers:
            ag_sheets = list(sheets.filter(agniveer=agniveer))
            if ag_sheets:
                row = build_tts_result_row(agniveer, ag_sheets)
                result_rows.append(row)
        
        return render(request, self.template_name, {
            'result_rows': result_rows,
            'trade': tts_trade,
            'department_name': 'TTS',
        })


class BattalionReportCardView(BattalionHeadMixin, View):
    """Battalion Head sees report cards for their battalion unit only"""
    template_name = 'reports/battalion_report_card.html'
    
    def get(self, request):
        user = request.user
        battalion_unit = user.battalion_unit
        
        # Filter agniveers by battalion unit
        if battalion_unit:
            agniveers = Agniveer.objects.filter(get_bn_desp_q('bn_desp', battalion_unit))
        else:
            battalion_units = get_bn_desp_list([choice[0] for choice in CustomUser.BATTALION_CHOICES])
            agniveers = Agniveer.objects.filter(bn_desp__in=battalion_units)
        
        sheets = EvaluationSheet.objects.filter(
            department='A', 
            is_locked=True,
            agniveer__in=agniveers
        ).select_related('agniveer').prefetch_related('marks')
        
        # Build result rows using existing Battalion result helper
        result_rows = []
        for agniveer in agniveers:
            ag_sheets = list(sheets.filter(agniveer=agniveer))
            if ag_sheets:
                row = build_battalion_result_row(agniveer, ag_sheets)
                result_rows.append(row)
        
        return render(request, self.template_name, {
            'result_rows': result_rows,
            'battalion_unit': battalion_unit,
            'department_name': 'Battalion',
        })


class CSReportCardView(CSHeadMixin, View):
    """CS Head sees report cards (loads report card without schema for now)"""
    template_name = 'reports/cs_report_card.html'
    
    def get(self, request):
        agniveers = Agniveer.objects.filter(evaluations__department='C').distinct()
        
        return render(request, self.template_name, {
            'agniveers': agniveers,
            'department_name': 'CS (Communication Systems)',
            'message': 'CS report card schema will be implemented soon.',
        })


class ClerkReportCardView(ClerkHeadMixin, View):
    """Clerk Head sees report cards (loads report card without schema for now)"""
    template_name = 'reports/clerk_report_card.html'
    
    def get(self, request):
        agniveers = Agniveer.objects.filter(trade__in=CLERK_TRADES)
        
        return render(request, self.template_name, {
            'agniveers': agniveers,
            'department_name': 'Clerk Department',
            'message': 'Clerk report card schema will be implemented soon.',
        })


# Report Card Detail View (Individual Agniveer)
class ReportCardDetailView(AnyStaffMixin, View):
    """Individual report card view with department-based access control"""
    template_name = 'reports/report_card_detail.html'
    
    def get(self, request, pk):
        agniveer = get_object_or_404(Agniveer, pk=pk)
        user = request.user
        
        # Access control: Check if user has permission to view this agniveer's report
        has_access = False
        is_department_view = False
        user_dept = None
        
        if user.is_commander or user.is_g_head:
            # Commander/G-Head can see ALL
            has_access = True
            
        elif user.is_department:
            user_dept = user.get_department_code()
            has_access = user_can_access_agniveer(user, agniveer, user_dept)
        
        if not has_access:
            return HttpResponse("You don't have permission to view this report card.", status=403)
        
        # Determine which evaluations to show (department-specific)
        if user.is_commander or user.is_g_head:
            evaluations = EvaluationSheet.objects.filter(agniveer=agniveer).prefetch_related('marks')
            departments_to_show = ['A', 'B', 'C', 'D']
            is_department_view = False
        else:
            user_dept = user.get_department_code()
            evaluations = scoped_sheets(
                EvaluationSheet.objects.filter(agniveer=agniveer).prefetch_related('marks'),
                user,
                user_dept,
            )
            departments_to_show = [user_dept]
            is_department_view = True
        
        # Build department evaluations data
        dept_evaluations = {}
        from evaluation.constants import get_dept_total_marks, get_overall_total_marks
        
        for dept in departments_to_show:
            dept_evals = evaluations.filter(department=dept)
            if dept_evals.exists():
                if dept == 'A':
                    d_row = build_department_result_row(agniveer, list(dept_evals), 'A')
                    total_marks = d_row.get('grand_total', 0.0)
                else:
                    total_marks = sum(e.get_total_marks() for e in dept_evals)
                dept_evaluations[dept] = {
                    'on_field': dept_evals.filter(category='on_field'),
                    'trade': dept_evals.filter(category='trade'),
                    'total_marks': total_marks,
                    'max_marks': get_dept_total_marks(dept),
                }
        
        # Calculate overall scores
        if is_department_view:
            result_row = build_department_result_row(agniveer, list(evaluations), user_dept)
            grand_total = result_row.get('grand_total', 0)
            max_total = result_row.get('max_total') or (120 if user_dept == 'A' else 40)
            percentage = result_row.get('percentage', 0)
            overall_pass = result_row.get('is_pass', False)
        else:
            # Compute summary for all 4 departments
            from evaluation.result_helpers import build_department_result_row, is_sheet_evaluated
            all_evaluations = list(EvaluationSheet.objects.filter(agniveer=agniveer).prefetch_related('marks'))
            
            trade = str(agniveer.trade or '').strip().upper()
            bn_raw = str(agniveer.bn_desp or '').strip().lower()
            if '1tb' in bn_raw:
                bn_name = '1tb'
            elif '2tb' in bn_raw or '2b' in bn_raw:
                bn_name = '2tb'
            else:
                bn_name = '1tb'
                
            dept_name_map = {
                'A': bn_name,
                'B': 'dmv department' if trade == 'DMV' else ('opem' if trade == 'OPEM' else 'tts department'),
                'C': 'CES DEpartment',
                'D': 'cts department',
            }
            
            all_dept_results = {}
            for d in ['A', 'B', 'C', 'D']:
                d_sheets = [s for s in all_evaluations if s.department == d]
                if d_sheets:
                    d_row = build_department_result_row(agniveer, d_sheets, d)
                    is_eval = any(is_sheet_evaluated(s) for s in d_sheets)
                    if is_eval:
                        all_dept_results[d] = {
                            'name': dept_name_map.get(d, d),
                            'grand_total': d_row.get('grand_total', 0),
                            'max_total': d_row.get('max_total') or (120 if d == 'A' else 40),
                        }
                    else:
                        all_dept_results[d] = {
                            'name': dept_name_map.get(d, d),
                            'grand_total': '—',
                            'max_total': (120 if d == 'A' else 40),
                        }
                else:
                    all_dept_results[d] = {
                        'name': dept_name_map.get(d, d),
                        'grand_total': '—',
                        'max_total': (120 if d == 'A' else 40),
                    }
            
            grand_total = 0.0
            max_total = 0.0
            for d, res in all_dept_results.items():
                if res['grand_total'] != '—':
                    grand_total += float(res['grand_total'])
                    max_total += float(res['max_total'])
            percentage = round((grand_total / max_total * 100), 2) if max_total else 0
            overall_pass = percentage >= 50
            
        passing_threshold = 40 if is_department_view and user_dept == 'A' else 50
        overall_pass = percentage >= passing_threshold
        
        # For TTS department, also build TTS result row
        tts_result_row = None
        if not is_department_view or user_dept == 'B':
            tts_sheets = evaluations.filter(department='B')
            if tts_sheets.exists():
                tts_result_row = build_tts_result_row(agniveer, list(tts_sheets))
        
        battalion_result_row = None
        if not is_department_view or user_dept == 'A':
            battalion_sheets = evaluations.filter(department='A')
            if battalion_sheets.exists():
                battalion_result_row = build_battalion_result_row(agniveer, list(battalion_sheets))
        
        return render(request, self.template_name, {
            'agniveer': agniveer,
            'dept_evaluations': dept_evaluations,
            'grand_total': grand_total,
            'max_total': max_total,
            'percentage': percentage,
            'overall_pass': overall_pass,
            'is_department_view': is_department_view,
            'user_department_name': DEPARTMENT_NAMES.get(user_dept, '') if is_department_view else None,
            'tts_result_row': tts_result_row,
            'battalion_result_row': battalion_result_row,
        })
class ExportAgniveersCSVView(AnyStaffMixin, View):
    def get(self, request):
        user = request.user
        queryset = Agniveer.objects.all()
        dept = user.get_department_code()

        if user.is_commander or user.is_g_head or user.is_registration_office:
            pass
        elif user.is_department:
            queryset = scoped_agniveers(queryset, user)
        else:
            queryset = Agniveer.objects.none()

        # TTS (Dept B) Specific Filtering
        if dept == 'B' and hasattr(user, 'tts_trade'):
            if user.tts_trade == 'DMV':
                queryset = queryset.filter(trade='DMV')
            elif user.tts_trade == 'OPEM':
                queryset = queryset.filter(trade='OPEM')
            elif user.tts_trade == 'OTHER':
                queryset = queryset.exclude(trade__in=['DMV', 'OPEM'])
                trade_param = request.GET.get('trade')
                if trade_param:
                    queryset = queryset.filter(trade=trade_param)

        # Clerk AV dashboard shows only Clerk trade Agniveers.
        if dept == 'D':
            queryset = queryset.filter(trade__in=['CLK', 'CLERK', 'Clerk', 'CLK_SD', 'CLK_IM'])

        # Battalion Level Isolation
        if user.is_battalion and user.battalion_unit:
            queryset = queryset.filter(get_bn_desp_q('bn_desp', user.battalion_unit))

        # Company/Platoon Level Isolation
        if not user.can_view_all:
            if hasattr(user, 'company') and user.company:
                queryset = queryset.filter(company=user.company)
            if hasattr(user, 'platoon') and user.platoon:
                queryset = queryset.filter(platoon=user.platoon)

        status = request.GET.get('status')
        search = request.GET.get('search')
        batch = request.GET.get('batch')
        batch_no = request.GET.get('batch_no')
        eval_filter = request.GET.get('eval_status')
        trade_param = request.GET.get('trade')
        battalion_param = request.GET.get('battalion')
        company_param = request.GET.get('company')
        platoon_param = request.GET.get('platoon')

        if status:
            queryset = queryset.filter(status=status)
        if batch:
            queryset = queryset.filter(batch__icontains=batch)
        if batch_no:
            queryset = queryset.filter(batch_no=batch_no)
        if search:
            queryset = queryset.filter(
                Q(enrollment_number__icontains=search) |
                Q(agniveer_no__icontains=search) |
                Q(name__icontains=search) |
                Q(father_name__icontains=search)
            )
        if trade_param:
            if not (dept == 'B' and hasattr(user, 'tts_trade') and user.tts_trade == 'OTHER'):
                queryset = queryset.filter(trade=trade_param)
        if battalion_param:
            queryset = queryset.filter(get_bn_desp_q('bn_desp', battalion_param))
        if company_param:
            queryset = queryset.filter(company=company_param)
        if platoon_param:
            queryset = queryset.filter(platoon=platoon_param)

        if eval_filter and (user.is_department or user.is_commander or user.is_g_head):
            from evaluation.result_helpers import is_sheet_evaluated
            if user.is_department:
                from evaluation.constants import get_dept_config
                config = get_dept_config(dept or 'A', user)
                test_types = [test[0] for test in config['test_types']]

                all_dept_sheets = EvaluationSheet.objects.filter(
                    department=dept,
                    test_type__in=test_types,
                ).prefetch_related('marks')
            else:
                all_dept_sheets = EvaluationSheet.objects.all().prefetch_related('marks')

            evaluated_ids = {s.agniveer_id for s in all_dept_sheets if is_sheet_evaluated(s)}

            if eval_filter == 'evaluated':
                queryset = queryset.filter(pk__in=evaluated_ids)
            elif eval_filter == 'not_evaluated':
                queryset = queryset.exclude(pk__in=evaluated_ids)

        sort_by = request.GET.get('sort_by', 'latest')
        if sort_by == 'oldest':
            agniveers = queryset.order_by('created_at', 'enrollment_number')
        else:
            agniveers = queryset.order_by('-created_at', '-enrollment_number')

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agniveers"
        st = _xl_styles(is_fail_theme=(eval_filter == 'not_evaluated'))

        # Build dynamic title based on filters
        title_parts = []
        
        # 1. Battalion
        battalion_val = battalion_param
        if not battalion_val and user.is_battalion and user.battalion_unit:
            battalion_val = user.battalion_unit
        if battalion_val:
            title_parts.append(f"Battalion - {battalion_val.lower()}")
            
        # 2. Company
        company_val = company_param
        if not company_val and not user.can_view_all and hasattr(user, 'company') and user.company:
            company_val = user.company
        if company_val:
            title_parts.append(f"Company {company_val}")
            
        # 3. Platoon
        platoon_val = platoon_param
        if not platoon_val and not user.can_view_all and hasattr(user, 'platoon') and user.platoon:
            platoon_val = user.platoon
        if platoon_val:
            title_parts.append(f"Platoon {platoon_val}")
            
        # 4. Trade
        trade_val = trade_param
        if not trade_val:
            if dept == 'B' and hasattr(user, 'tts_trade') and user.tts_trade != 'OTHER':
                trade_val = user.tts_trade
            elif dept == 'D':
                trade_val = 'clerk'
            elif dept == 'C':
                trade_val = 'ces'
        if trade_val:
            title_parts.append(trade_val.lower())
            
        # 5. Eval Status
        eval_status_val = eval_filter
        if eval_status_val:
            if eval_status_val == 'evaluated':
                title_parts.append("evaluated")
            elif eval_status_val == 'not_evaluated':
                title_parts.append("non evaluated")

        if title_parts:
            title_text = " - ".join(title_parts).title() + " Agniveers"
        else:
            title_text = "All Agniveers"

        headers = ['Agniveer No', 'Name', 'Rank', 'B/desp', 'Company', 'Platoon', 'Trade']
        num_cols = len(headers)

        _xl_write_title(ws, title_text, num_cols, st)
        _xl_write_headers(ws, headers, st)

        NAME_COL = 2
        for index, a in enumerate(agniveers, 1):
            values = [
                a.agniveer_no or a.enrollment_number,
                a.get_full_name(),
                getattr(a, 'rank', '') or '',
                a.bn_desp or '',
                a.company or '',
                a.platoon or '',
                a.trade or '',
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=index + 2, column=col, value=value)
                cell.font = st['data_font']
                cell.border = st['border']
                cell.alignment = st['left'] if col == NAME_COL else st['center']
            ws.row_dimensions[index + 2].height = 20

        for col_idx in range(1, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            vals = [ws.cell(row=r, column=col_idx).value for r in range(2, len(agniveers) + 3)]
            max_len = max((len(str(v or '')) for v in vals), default=12)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(num_cols)}{max(len(agniveers) + 2, 2)}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        import re
        safe_filename = re.sub(r'[^a-zA-Z0-9_\-]', '_', title_text.strip().lower())
        safe_filename = f"{safe_filename}.xlsx"

        log_action(user, 'EXPORT', f'Exported Agniveers Excel: {title_text}', request)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        return response


class ExportEvaluationsCSVView(AnyStaffMixin, View):
    def get(self, request):
        user = request.user
        sheets = EvaluationSheet.objects.filter(is_locked=True).select_related('agniveer').prefetch_related('marks')

        if user.is_department:
            sheets = scoped_sheets(sheets, user)
        else:
            sheets = EvaluationSheet.objects.none()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="evaluations.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Enrollment No', 'Name', 'Department', 'Category', 'Test Type',
            'NCO Marks', 'JCO Marks', 'Officer Marks', 'Total', 'Percentage', 'Result'
        ])

        for sheet in sheets:
            writer.writerow([
                sheet.agniveer.enrollment_number,
                sheet.agniveer.get_full_name(),
                DEPARTMENT_NAMES.get(sheet.department, sheet.department),
                sheet.get_category_display(),
                sheet.get_test_type_display(),
                sheet.get_nco_marks(),
                sheet.get_jco_marks(),
                sheet.get_officer_marks(),
                sheet.get_total_marks(),
                f'{sheet.get_percentage()}%',
                'Pass' if sheet.is_pass() else 'Fail',
            ])

        log_action(user, 'EXPORT', 'Exported Evaluations CSV', request)
        return response


class ExportExcelView(AnyStaffMixin, View):
    def get(self, request):
        try:
            import openpyxl
        except ImportError:
            return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

        user = request.user
        sheets = EvaluationSheet.objects.filter(is_locked=True).select_related('agniveer').prefetch_related('marks')

        if user.is_department:
            sheets = scoped_sheets(sheets, user)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Report"
        st = _xl_styles()

        headers = [
            'S.No', 'Enrollment No', 'Army No', 'Rank', 'Trade', 'Name', 'Unit', 'Department', 'Category',
            'Test Type', 'Total Marks', 'Percentage', 'Result'
        ]
        _xl_write_title(ws, "ARMY EVALUATION PORTAL - EVALUATION REPORT", len(headers), st)
        _xl_write_headers(ws, headers, st)

        NAME_COL = 6
        RESULT_COL = 13

        for row_idx, sheet in enumerate(sheets, 1):
            is_pass = sheet.is_pass()
            data = [
                row_idx,
                sheet.agniveer.enrollment_number,
                sheet.agniveer.agniveer_no or sheet.agniveer.enrollment_number,
                getattr(sheet.agniveer, 'rank', '') or '',
                sheet.agniveer.trade or '',
                sheet.agniveer.get_full_name(),
                sheet.agniveer.bn_desp or '',
                DEPARTMENT_NAMES.get(sheet.department, sheet.department),
                sheet.get_category_display(),
                sheet.get_test_type_display(),
                sheet.get_total_marks(),
                f"{sheet.get_percentage()}%",
                'PASS' if is_pass else 'FAIL',
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx + 2, column=col)
                _xl_style_data_cell(cell, value, is_pass,
                                    is_result_col=(col == RESULT_COL),
                                    is_name_col=(col == NAME_COL), styles=st)
            ws.row_dimensions[row_idx + 2].height = 20

        # Column widths
        col_widths = [6, 18, 15, 10, 12, 24, 12, 16, 16, 18, 14, 12, 10]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(len(sheets) + 2, 2)}"

        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_action(user, 'EXPORT', 'Exported Excel Report', request)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="evaluation_report.xlsx"'
        return response


class ExportTestTypeExcelView(AnyStaffMixin, View):
    """Export locked evaluation sheets for a specific department and test type as Excel."""
    def get(self, request, dept, test_type):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

        user = request.user

        # Restrict dept to user's dept unless commander/g_head
        if not (user.is_commander or user.is_g_head):
            user_dept = user.get_department_code()
            if dept != user_dept:
                return HttpResponse("Not authorized to export for this department", status=403)

        sheets = EvaluationSheet.objects.filter(is_locked=True, test_type=test_type, department=dept).select_related('agniveer').prefetch_related('marks')

        # Optional status filter: pass|fail
        status = request.GET.get('status')
        if status in ('pass', 'fail'):
            wanted = True if status == 'pass' else False
            sheets = [s for s in sheets if s.is_pass() == wanted]

        # Further restrict to scoped sheets for department users
        if user.is_department and not (user.is_commander or user.is_g_head):
            sheets = scoped_sheets(sheets, user, dept)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{dept}_{test_type}"[:31]
        st = _xl_styles(is_fail_theme=(status == 'fail'))

        headers = ['S.No', 'Enrollment No', 'Army No', 'Name', 'Trade', 'Unit', 'Category', 'Test Type', 'Total Marks', 'Percentage', 'Result']

        # Title row
        dept_label = DEPARTMENT_NAMES.get(dept, dept)
        _xl_write_title(ws, f"ARMY EVALUATION PORTAL - {dept_label.upper()} - {test_type.replace('_', ' ')} ({(status or 'ALL').upper()} LIST)", len(headers), st)
        _xl_write_headers(ws, headers, st)

        NAME_COL = 4
        RESULT_COL = 11

        for idx, sheet in enumerate(sheets, 1):
            is_pass = sheet.is_pass()
            row = [
                idx,
                sheet.agniveer.enrollment_number,
                sheet.agniveer.agniveer_no or sheet.agniveer.enrollment_number,
                sheet.agniveer.get_full_name(),
                sheet.agniveer.trade or '',
                sheet.agniveer.bn_desp or '',
                sheet.get_category_display(),
                sheet.get_test_type_display(),
                sheet.get_total_marks(),
                f"{sheet.get_percentage()}%",
                'PASS' if is_pass else 'FAIL',
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=idx + 2, column=col)
                _xl_style_data_cell(cell, val, is_pass,
                                    is_result_col=(col == RESULT_COL),
                                    is_name_col=(col == NAME_COL), styles=st)
            ws.row_dimensions[idx + 2].height = 20

        ws.freeze_panes = 'A3'
        col_widths = [6, 18, 15, 24, 12, 12, 16, 18, 14, 12, 10]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(len(sheets) + 2, 2)}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{DEPARTMENT_NAMES.get(dept,dept)}_{test_type}_report.xlsx"
        if status:
            filename = f"{DEPARTMENT_NAMES.get(dept,dept)}_{test_type}_{status}_report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        log_action(user, 'EXPORT', f'Exported {dept} {test_type} Excel', request)
        return response


class TestTypeResultsView(AnyStaffMixin, View):
    """Show the locked evaluation sheets for a given department and test type. Renders into the reports dashboard template (no new template file)."""

    def get(self, request, dept, test_type):
        user = request.user

        # Authorization: department users only for their own dept unless commander/g_head
        if user.is_department and not (user.is_commander or user.is_g_head):
            if dept != user.get_department_code():
                return HttpResponse("Not authorized to view this department results", status=403)

        sheets_qs = EvaluationSheet.objects.filter(is_locked=True, department=dept, test_type=test_type).select_related('agniveer').prefetch_related('marks')
        if user.is_department and not (user.is_commander or user.is_g_head):
            sheets_qs = scoped_sheets(sheets_qs, user, dept)

        rows = []
        for s in sheets_qs:
            rows.append({
                'agniveer': s.agniveer,
                'nco': s.get_nco_marks(),
                'jco': s.get_jco_marks(),
                'officer': s.get_officer_marks(),
                'total': s.get_total_marks(),
                'percentage': s.get_percentage(),
                'is_pass': s.is_pass(),
            })

        # Counts
        passed = sum(1 for r in rows if r['is_pass'])
        failed = sum(1 for r in rows if not r['is_pass'])

        # Reuse ReportDashboardView context by calling it to get base context
        base_view = ReportDashboardView()
        base_request = request
        base_ctx = base_view.get(request).context_data if hasattr(base_view.get(request), 'context_data') else {}

        # Construct context and render the same dashboard template with the results block
        context = base_ctx
        context.update({
            'test_results_rows': rows,
            'test_results_passed': passed,
            'test_results_failed': failed,
            'test_results_dept': dept,
            'test_results_test_type': test_type,
        })

        return render(request, 'reports/report_dashboard.html', context)


class DeptTestResultsView(AnyStaffMixin, View):
    """Render test-type results into the department dashboard template for inline viewing."""

    def get(self, request, dept, test_type):
        user = request.user
        # Authorization
        if user.is_department and not (user.is_commander or user.is_g_head):
            if dept != user.get_department_code():
                return HttpResponse("Not authorized to view this department results", status=403)

        sheets_qs = EvaluationSheet.objects.filter(is_locked=True, department=dept, test_type=test_type).select_related('agniveer').prefetch_related('marks')
        if user.is_department and not (user.is_commander or user.is_g_head):
            sheets_qs = scoped_sheets(sheets_qs, user, dept)

        rows = []
        for s in sheets_qs:
            rows.append({
                'agniveer': s.agniveer,
                'nco': s.get_nco_marks(),
                'jco': s.get_jco_marks(),
                'officer': s.get_officer_marks(),
                'total': s.get_total_marks(),
                'percentage': s.get_percentage(),
                'is_pass': s.is_pass(),
            })

        passed = sum(1 for r in rows if r['is_pass'])
        failed = sum(1 for r in rows if not r['is_pass'])

        # Build department dashboard context by delegating to core view logic
        from core.views import get_all_test_types_for_dept
        # Minimal context: include test results and available test types
        context = {
            'dept': dept,
            'test_results_rows': rows,
            'test_results_passed': passed,
            'test_results_failed': failed,
            'test_results_dept': dept,
            'test_results_test_type': test_type,
            'test_types': get_all_test_types_for_dept(dept),
        }

        return render(request, 'core/department_dashboard_advanced.html', context)



class DeptTestResultsJsonView(AnyStaffMixin, View):
    """Return JSON list of pass/fail results for a department + test_type for inline dashboard loading."""

    def get(self, request, dept, test_type):
        from django.http import JsonResponse
        user = request.user

        if user.is_department and not (user.is_commander or user.is_g_head):
            if dept != user.get_department_code():
                return JsonResponse({'error': 'Not authorized'}, status=403)

        sheets_qs = (
            EvaluationSheet.objects
            .filter(is_locked=True, department=dept, test_type=test_type)
            .select_related('agniveer')
            .prefetch_related('marks')
        )
        if user.is_department and not (user.is_commander or user.is_g_head):
            sheets_qs = list(scoped_sheets(sheets_qs, user, dept))

        rows = []
        for s in sheets_qs:
            is_pass = s.is_pass()
            rows.append({
                'id': s.agniveer.pk,
                'army_no': s.agniveer.agniveer_no or s.agniveer.enrollment_number,
                'name': s.agniveer.get_full_name(),
                'batch': s.agniveer.batch or '—',
                'total': s.get_total_marks(),
                'max': s.get_max_marks(),
                'percentage': s.get_percentage(),
                'is_pass': is_pass,
            })

        passed = sum(1 for r in rows if r['is_pass'])
        failed = len(rows) - passed

        return JsonResponse({
            'dept': dept,
            'test_type': test_type,
            'total': len(rows),
            'passed': passed,
            'failed': failed,
            'rows': rows,
        })



def _xl_styles(is_fail_theme=False):
    """
    Returns a dict of shared openpyxl style objects used uniformly across all Excel exports.
    If is_fail_theme is True, returns a red-themed palette. Otherwise, returns a green-themed palette.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if is_fail_theme:
        BRAND_COLOR   = "7F1D1D"   # Dark red
        TITLE_BG      = "FEF2F2"   # Very light red for title row
        HEADER_BG     = "7F1D1D"   # Same dark red for header row
        HEADER_BORDER = "FCA5A5"   # Red-tinted border colour
    else:
        BRAND_COLOR   = "1B4332"   # Dark army green
        TITLE_BG      = "E8F5E9"   # Very light green for title row
        HEADER_BG     = "1B4332"   # Same dark green for header row
        HEADER_BORDER = "A5D6A7"   # Green-tinted border colour

    PASS_ROW_BG   = "DDF4E8"   # Soft green row fill
    PASS_RESULT   = "1B5E20"   # Deep green text for PASS cell
    FAIL_ROW_BG   = "FCE4E4"   # Soft red row fill
    FAIL_RESULT   = "B71C1C"   # Deep red text for FAIL cell

    header_fill  = PatternFill(start_color=HEADER_BG,   end_color=HEADER_BG,   fill_type="solid")
    title_fill   = PatternFill(start_color=TITLE_BG,    end_color=TITLE_BG,    fill_type="solid")
    pass_fill    = PatternFill(start_color=PASS_ROW_BG, end_color=PASS_ROW_BG, fill_type="solid")
    fail_fill    = PatternFill(start_color=FAIL_ROW_BG, end_color=FAIL_ROW_BG, fill_type="solid")

    header_font      = Font(color="FFFFFF", bold=True, size=11, name='Calibri')
    title_font       = Font(color=BRAND_COLOR, bold=True, size=13, name='Calibri')
    pass_result_font = Font(color=PASS_RESULT, bold=True, size=10, name='Calibri')
    fail_result_font = Font(color=FAIL_RESULT, bold=True, size=10, name='Calibri')
    data_font        = Font(color="1A1A1A", size=10, name='Calibri')

    thin_border_side = Side(style='thin', color=HEADER_BORDER)
    border           = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    return {
        'header_fill': header_fill,
        'title_fill':  title_fill,
        'pass_fill':   pass_fill,
        'fail_fill':   fail_fill,
        'header_font':      header_font,
        'title_font':       title_font,
        'pass_result_font': pass_result_font,
        'fail_result_font': fail_result_font,
        'data_font':        data_font,
        'border':  border,
        'center':  center,
        'left':    left,
        'brand_color': BRAND_COLOR,
    }


def _xl_write_title(ws, title_text, num_cols, styles):
    """Writes a merged, styled title in row 1."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font      = styles['title_font']
    cell.fill      = styles['title_fill']
    cell.alignment = styles['center']
    ws.row_dimensions[1].height = 34


def _xl_write_headers(ws, headers, styles, start_row=2):
    """Writes the dark-green header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill      = styles['header_fill']
        cell.font      = styles['header_font']
        cell.alignment = styles['center']
        cell.border    = styles['border']
    ws.row_dimensions[start_row].height = 28


def _xl_style_data_cell(cell, value, is_pass, is_result_col, is_name_col, styles):
    """Applies row fill, fonts, and special result-cell colouring."""
    cell.value     = value
    cell.fill      = styles['pass_fill'] if is_pass else styles['fail_fill']
    cell.alignment = styles['left'] if is_name_col else styles['center']
    cell.border    = styles['border']
    if is_result_col:
        cell.font = styles['pass_result_font'] if is_pass else styles['fail_result_font']
    else:
        cell.font = styles['data_font']


class ExportDashboardResultsExcelView(AnyStaffMixin, View):
    def _export_battalion_results(self, request, status_filter, sub_dept=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.db.models import Q
        from evaluation.models import Marks, EvaluationSheet
        from evaluation.result_helpers import build_battalion_result_row
        from departments.models import Agniveer

        from evaluation.result_helpers import is_sheet_evaluated
        user = request.user
        is_pass_filter = status_filter == 'pass'

        agniveers = Agniveer.objects.all()
        sheets = EvaluationSheet.objects.all().select_related('agniveer').prefetch_related('marks')

        unit_filter = sub_dept
        if not unit_filter and user.is_battalion and user.battalion_unit:
            unit_filter = user.battalion_unit

        if unit_filter and unit_filter != 'all':
            agniveers = agniveers.filter(get_bn_desp_q('bn_desp', unit_filter))
            sheets = sheets.filter(get_bn_desp_q('agniveer__bn_desp', unit_filter))

        from collections import defaultdict
        sheets_by_agniveer = defaultdict(list)
        for s in sheets:
            if is_sheet_evaluated(s):
                sheets_by_agniveer[s.agniveer_id].append(s)

        rows = []
        for agniveer in agniveers.order_by('agniveer_no', 'enrollment_number'):
            ag_sheets = sheets_by_agniveer.get(agniveer.id, [])
            if not ag_sheets:
                continue
            row = build_battalion_result_row(agniveer, ag_sheets)
            if status_filter == 'evaluated':
                rows.append(row)   # include both pass and fail
            elif row['is_pass'] == is_pass_filter:
                rows.append(row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"BN {status_filter.upper()}"
        st = _xl_styles(is_fail_theme=(status_filter == 'fail'))

        headers = [
            'ARMY NO', 'RANK', 'TRADE', 'NAME',
            'COMMON MIL KNOWLEDGE (20)', 'BASIC TACTICS (CES) (40)',
            'TRADE PROFICIENCY (BTT) (40)', 'WPN & EQPT HANDLING (20)',
            'TOTAL (120)', 'ROUND FIGURE (120)', '%', 'RESULT'
        ]
        if status_filter != 'fail':
            headers.append('GRADING')

        unit_label = sub_dept.upper() if sub_dept and sub_dept != 'all' else 'ALL BATTALIONS'
        _xl_write_title(ws, f"BATTALION SCREENING – {unit_label} – {status_filter.upper()} LIST", len(headers), st)
        _xl_write_headers(ws, headers, st)

        NAME_COL   = 4
        RESULT_COL = 12

        for index, row in enumerate(rows, 1):
            is_pass = row['is_pass']
            values = [
                row['army_no'], row['rank'], row['trade'], row['name'],
                row['cmk_20'], row['basic_tac_40'], row['trade_prof_40'], row['wpn_handling_20'],
                row['total_120'], row['round_figure_120'],
                f"{row['percentage']}%" if row['percentage'] else '0%',
                'PASS' if is_pass else 'FAIL',
            ]
            if status_filter != 'fail':
                values.append(row.get('grading', '—'))
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=index + 2, column=col)
                _xl_style_data_cell(cell, value, is_pass,
                                    is_result_col=(col == RESULT_COL),
                                    is_name_col=(col == NAME_COL), styles=st)
            ws.row_dimensions[index + 2].height = 20

        widths = [16, 10, 14, 28, 28, 28, 28, 26, 14, 18, 10, 10]
        if status_filter != 'fail':
            widths.append(14)
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(len(rows) + 2, 2)}"


        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from logs.utils import log_action
        log_action(user, 'EXPORT', f'Exported BN {status_filter} dashboard results Excel', request)
        from django.http import HttpResponse
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="bn_{status_filter}_results.xlsx"'
        return response

    def _export_tts_results(self, request, status_filter, sub_dept=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.db.models import Q
        from evaluation.models import Marks, EvaluationSheet
        from evaluation.result_helpers import build_tts_result_row, _marks_from_sheet, _num, get_grade
        from departments.models import Agniveer

        user = request.user
        is_pass_filter = status_filter == 'pass'
        from evaluation.result_helpers import is_sheet_evaluated
        agniveers = Agniveer.objects.exclude(trade__in=CLERK_TRADES)
        agniveers = scoped_agniveers(agniveers, user, 'B')
        sheets = EvaluationSheet.objects.filter(
            department='B'
        ).select_related('agniveer').prefetch_related('marks')

        trade_filter = sub_dept
        if not trade_filter and getattr(user, 'tts_trade', None):
            trade_filter = user.tts_trade

        if trade_filter and trade_filter != 'all':
            if trade_filter == 'DMV':
                agniveers = agniveers.filter(trade='DMV')
                sheets = sheets.filter(agniveer__trade='DMV')
            elif trade_filter == 'OPEM':
                agniveers = agniveers.filter(trade='OPEM')
                sheets = sheets.filter(agniveer__trade='OPEM')
            elif trade_filter == 'OTHER':
                agniveers = agniveers.exclude(trade__in=['DMV', 'OPEM'] + CLERK_TRADES)
                sheets = sheets.exclude(agniveer__trade__in=['DMV', 'OPEM'] + CLERK_TRADES)
        else:
            agniveers = agniveers.exclude(trade__in=CLERK_TRADES)
            sheets = sheets.exclude(agniveer__trade__in=CLERK_TRADES)

        from collections import defaultdict
        sheets_by_agniveer = defaultdict(list)
        for s in sheets:
            if is_sheet_evaluated(s):
                sheets_by_agniveer[s.agniveer_id].append(s)

        rows = []
        for agniveer in agniveers.order_by('agniveer_no', 'enrollment_number'):
            ag_sheets = sheets_by_agniveer.get(agniveer.id, [])
            if not ag_sheets:
                continue
            
            row = build_tts_result_row(agniveer, ag_sheets)
            if status_filter == 'evaluated':
                rows.append(row)
            elif row['is_pass'] == is_pass_filter:
                rows.append(row)

        wb = openpyxl.Workbook()
        ws = wb.active
        st = _xl_styles(is_fail_theme=(status_filter == 'fail'))

        if trade_filter == 'DMV':
            ws.title = f"DMV {status_filter.upper()}"
            title_text = "FINAL RESULT SHEET OF DMV"
            headers = [
                'ARMY NO', 'RANK', 'TRADE', 'NAME', 'UNIT',
                'Online Test (100)', 'Practical Test (50)', 'Driving Test (50)',
                'Total (200)', '% Age', 'RESULT'
            ]
            if status_filter != 'fail':
                headers.append('Grading')
            headers.extend(['Convert 40 Marks', 'REMARKS'])
            RESULT_COL = 11
        elif trade_filter == 'OPEM':
            ws.title = f"OPEM {status_filter.upper()}"
            title_text = "FINAL RESULT SHEET OF OPEM"
            headers = [
                'ARMY NO', 'RANK', 'TRADE', 'NAME', 'UNIT',
                'Written Test (100)', 'Practical Test (50)', 'Maintenance Test (50)',
                'Total (200)', '% Age', 'RESULT'
            ]
            if status_filter != 'fail':
                headers.append('Grading')
            headers.extend(['Convert 40 Marks', 'REMARKS'])
            RESULT_COL = 11
        else:
            ws.title = f"TTS {status_filter.upper()}"
            title_text = "SCREEN BOARD OF AGNIVEER (TTS RESULT)"
            headers = [
                'ARMY NO', 'RANK', 'TRADE', 'NAME', 'UNIT',
                'Mid Term Test (50)', 'Convert In 10 Mks (Mid Term)',
                'Online Test (100)', 'Convert In 15 Mks',
                'Job (40)', 'Convert In 05 Mks',
                'Practical (60)', 'Convert In 10 Mks (Practical)',
                'Grand Total (40)', '% Age', 'RESULT'
            ]
            if status_filter != 'fail':
                headers.append('Grading')
            RESULT_COL = 16

        NAME_COL = 4
        _xl_write_title(ws, title_text, len(headers), st)
        _xl_write_headers(ws, headers, st)

        for index, r in enumerate(rows, 1):
            is_pass = r['is_pass']
            if trade_filter == 'DMV':
                values = [
                    r['army_no'], r['rank'], r['trade'], r['name'], r['unit'],
                    r['online'], r['practical'], r['job'],
                    r['total_200'], f"{r['percentage']}%" if r['percentage'] else '0%',
                    'PASS' if is_pass else 'FAIL',
                ]
                if status_filter != 'fail':
                    values.append(r['grading'])
                values.extend([r['grand_total'], r['remarks']])
            elif trade_filter == 'OPEM':
                values = [
                    r['army_no'], r['rank'], r['trade'], r['name'], r['unit'],
                    r['online'], r['practical'], r['job'],
                    r['total_200'], f"{r['percentage']}%" if r['percentage'] else '0%',
                    'PASS' if is_pass else 'FAIL',
                ]
                if status_filter != 'fail':
                    values.append(r['grading'])
                values.extend([r['grand_total'], r['remarks']])
            else:
                values = [
                    r['army_no'], r['rank'], r['trade'], r['name'], r['unit'],
                    r['mid_term'], r['mid_term_conv'],
                    r['online'], r['online_conv'],
                    r['job'], r['job_conv'],
                    r['practical'], r['practical_conv'],
                    r['grand_total'], f"{r['percentage']}%" if r['percentage'] else '0%',
                    'PASS' if is_pass else 'FAIL',
                ]
                if status_filter != 'fail':
                    values.append(r['grading'])
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=index + 2, column=col)
                _xl_style_data_cell(cell, value, is_pass,
                                    is_result_col=(col == RESULT_COL),
                                    is_name_col=(col == NAME_COL), styles=st)
            ws.row_dimensions[index + 2].height = 20

        for col_idx in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            vals = [ws.cell(row=r, column=col_idx).value for r in range(2, len(rows) + 3)]
            max_len = max((len(str(v or '')) for v in vals), default=10)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(len(rows)+2,2)}"

        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_action(user, 'EXPORT', f'Exported TTS {status_filter} dashboard results Excel', request)
        from django.http import HttpResponse
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tts_{status_filter}_results.xlsx"'
        return response

    def _export_cs_results(self, request, status_filter, sub_dept=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.db.models import Q
        from evaluation.models import Marks, EvaluationSheet
        from evaluation.result_helpers import build_cs_result_row, _marks_from_sheet, _num, get_grade
        from departments.models import Agniveer
        import io

        user = request.user
        is_pass_filter = status_filter == 'pass'
        from evaluation.result_helpers import is_sheet_evaluated
        agniveers = Agniveer.objects.all()
        agniveers = scoped_agniveers(agniveers, user, 'C')
        
        sheets = EvaluationSheet.objects.filter(
            department='C'
        ).select_related('agniveer').prefetch_related('marks')

        if sub_dept and sub_dept != 'all':
            agniveers = agniveers.filter(get_bn_desp_q('bn_desp', sub_dept))
            sheets = sheets.filter(get_bn_desp_q('agniveer__bn_desp', sub_dept))

        cs_final_rows = []
        cs_clerk_rows = []

        from evaluation.constants import CS_CLERK_RESULT_TRADES

        from collections import defaultdict
        sheets_by_agniveer = defaultdict(list)
        for s in sheets:
            if is_sheet_evaluated(s):
                sheets_by_agniveer[s.agniveer_id].append(s)

        for agniveer in agniveers.order_by('agniveer_no', 'enrollment_number'):
            ag_sheets = sheets_by_agniveer.get(agniveer.id, [])
            if not ag_sheets:
                continue
            
            sheet_map = {s.test_type: s for s in ag_sheets}
            
            if agniveer.trade in CS_CLERK_RESULT_TRADES:
                result_sheet = sheet_map.get('CS_CLERK_RESULT') or sheet_map.get('CS_ASSESSMENT')
                if not result_sheet:
                    continue
                marks = _marks_from_sheet(result_sheet)
                online = _num(marks.get('Online (20)'))
                prac = _num(marks.get('TPrac (20)'))
                total = _num(marks.get('Total (40)')) or (online + prac)
                percentage = round((total / 40) * 100, 2) if total else 0
                is_pass = percentage >= 50
                
                if status_filter == 'evaluated' or is_pass == is_pass_filter:
                    cs_clerk_rows.append({
                        'rank': getattr(agniveer, 'rank', '') or '',
                        'trade': agniveer.trade or '',
                        'name': agniveer.get_full_name(),
                        'pl': agniveer.platoon or '',
                        'bn': agniveer.bn_desp or '',
                        'online': online,
                        'prac': prac,
                        'total': total,
                        'remarks': result_sheet.remarks if result_sheet.remarks else ''
                    })
            else:
                result_sheet = sheet_map.get('CS_RESULT') or sheet_map.get('CS_ASSESSMENT')
                if not result_sheet:
                    continue
                marks = _marks_from_sheet(result_sheet)
                
                toet_i = _num(marks.get('TOET-I (25)'))
                toet_ii = _num(marks.get('TOET-II (25)'))
                toet_total = _num(marks.get('TOTAL TOET (50)')) or (toet_i + toet_ii)
                toet_25 = _num(marks.get('25% OF TOET (25)')) or round(toet_total * 0.5, 2)
                fe_online = _num(marks.get('FE Online Exam (50)'))
                fe_prac = _num(marks.get('FE Prac (20)'))
                fe_total = _num(marks.get('FE Total (70)')) or (fe_online + fe_prac)
                br_online = _num(marks.get('BR Online Exam (40)'))
                br_prac = _num(marks.get('BR Prac (25)'))
                br_total = _num(marks.get('BR Total (65)')) or (br_online + br_prac)
                total_160 = _num(marks.get('TOTAL (160)')) or (toet_total + fe_total + br_total)
                converted_40 = _num(marks.get('CONVERTED TO 40')) or round(total_160 * 0.25, 2)
                percentage = round((converted_40 / 40) * 100, 2) if converted_40 else 0
                is_pass = percentage >= 50
                
                if status_filter == 'evaluated' or is_pass == is_pass_filter:
                    cs_final_rows.append({
                        'army_no': agniveer.agniveer_no or agniveer.enrollment_number,
                        'rank': getattr(agniveer, 'rank', '') or '',
                        'trade': agniveer.trade or '',
                        'name': agniveer.get_full_name(),
                        'unit': agniveer.bn_desp or '',
                        'toet_i': toet_i,
                        'toet_ii': toet_ii,
                        'toet_total': toet_total,
                        'toet_25': toet_25,
                        'fe_online': fe_online,
                        'fe_prac': fe_prac,
                        'fe_total': fe_total,
                        'br_online': br_online,
                        'br_prac': br_prac,
                        'br_total': br_total,
                        'total_160': total_160,
                        'converted_40': converted_40,
                        'remarks': result_sheet.remarks if result_sheet.remarks else ''
                    })

        format_param = request.GET.get('format')  # 'final' or 'clerk'
        wb = openpyxl.Workbook()
        st = _xl_styles(is_fail_theme=(status_filter == 'fail'))

        # ── Main (CES Final) Sheet ──────────────────────────────────────────
        if not format_param or format_param == 'final':
            ws1 = wb.active
            ws1.title = "CES Final"

            headers1 = [
                'ARMY NO', 'RANK', 'TRADE', 'NAME', 'UNIT',
                'TOET-I (25)', 'TOET-II (25)', 'TOET TOTAL (50)', '25% OF TOET (25)',
                'FE Online Exam (50)', 'FE Prac (20)', 'FE Total (70)',
                'BR Online Exam (40)', 'BR Prac (25)', 'BR Total (65)',
                'TOTAL (160)', 'CONVERTED TO 40', 'RESULT', 'REMARKS'
            ]
            RESULT_COL1 = 18
            NAME_COL = 4

            _xl_write_title(ws1, f"FINAL RESULT OF CES – {status_filter.upper()} LIST", len(headers1), st)
            _xl_write_headers(ws1, headers1, st)

            for index, r in enumerate(cs_final_rows, 1):
                is_pass = status_filter == 'pass'
                values = [
                    r['army_no'], r['rank'], r['trade'], r['name'], r['unit'],
                    r['toet_i'], r['toet_ii'], r['toet_total'], r['toet_25'],
                    r['fe_online'], r['fe_prac'], r['fe_total'],
                    r['br_online'], r['br_prac'], r['br_total'],
                    r['total_160'], r['converted_40'],
                    'PASS' if is_pass else 'FAIL', r['remarks']
                ]
                for col, value in enumerate(values, 1):
                    cell = ws1.cell(row=index + 2, column=col)
                    _xl_style_data_cell(cell, value, is_pass,
                                        is_result_col=(col == RESULT_COL1),
                                        is_name_col=(col == NAME_COL), styles=st)
                ws1.row_dimensions[index + 2].height = 20

            for col_idx in range(1, len(headers1) + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                vals = [ws1.cell(row=r, column=col_idx).value for r in range(2, len(cs_final_rows) + 3)]
                ws1.column_dimensions[col_letter].width = max((len(str(v or '')) for v in vals), default=10) + 4
            ws1.freeze_panes = 'A3'
            ws1.auto_filter.ref = f"A2:S{max(len(cs_final_rows)+2, 2)}"

        # ── Clerk Sheet ─────────────────────────────────────────────────────
        if not format_param or format_param == 'clerk':
            ws2 = wb.create_sheet(title="CES Clerk Final") if not format_param else wb.active
            if format_param:
                ws2.title = "CES Clerk Final"

            headers2 = [
                'RANK', 'TRADE', 'NAME', 'PL', 'BN',
                'Online (20)', 'Prac (20)', 'Total (40)', 'RESULT', 'REMARKS'
            ]
            RESULT_COL2 = 9

            _xl_write_title(ws2, f"FINAL RESULT OF CES (CLERK) – {status_filter.upper()} LIST", len(headers2), st)
            _xl_write_headers(ws2, headers2, st)

            for index, r in enumerate(cs_clerk_rows, 1):
                is_pass = status_filter == 'pass'
                values = [
                    r['rank'], r['trade'], r['name'], r['pl'], r['bn'],
                    r['online'], r['prac'], r['total'],
                    'PASS' if is_pass else 'FAIL', r['remarks']
                ]
                for col, value in enumerate(values, 1):
                    cell = ws2.cell(row=index + 2, column=col)
                    _xl_style_data_cell(cell, value, is_pass,
                                        is_result_col=(col == RESULT_COL2),
                                        is_name_col=(col == 3), styles=st)
                ws2.row_dimensions[index + 2].height = 20

            for col_idx in range(1, len(headers2) + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                vals = [ws2.cell(row=r, column=col_idx).value for r in range(2, len(cs_clerk_rows) + 3)]
                ws2.column_dimensions[col_letter].width = max((len(str(v or '')) for v in vals), default=10) + 4
            ws2.freeze_panes = 'A3'
            ws2.auto_filter.ref = f"A2:J{max(len(cs_clerk_rows)+2, 2)}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_action(user, 'EXPORT', f'Exported CS {status_filter} dashboard results Excel', request)
        suffix = f"_{format_param}" if format_param else ""
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cs_{status_filter}_results{suffix}.xlsx"'
        return response

    def _export_clerk_results(self, request, status_filter, sub_dept=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.db.models import Q
        from evaluation.models import Marks, EvaluationSheet
        from evaluation.result_helpers import build_clerk_result_row, _marks_from_sheet, _num, get_grade
        from departments.models import Agniveer
        import io

        user = request.user
        is_pass_filter = status_filter == 'pass'
        from evaluation.result_helpers import is_sheet_evaluated
        agniveers = Agniveer.objects.all()
        agniveers = scoped_agniveers(agniveers, user, 'D')
        
        sheets = EvaluationSheet.objects.filter(
            department='D'
        ).select_related('agniveer').prefetch_related('marks')

        if sub_dept and sub_dept != 'all':
            agniveers = agniveers.filter(get_bn_desp_q('bn_desp', sub_dept))
            sheets = sheets.filter(get_bn_desp_q('agniveer__bn_desp', sub_dept))

        from collections import defaultdict
        sheets_by_agniveer = defaultdict(list)
        for s in sheets:
            if is_sheet_evaluated(s):
                sheets_by_agniveer[s.agniveer_id].append(s)

        rows = []
        for agniveer in agniveers.order_by('agniveer_no', 'enrollment_number'):
            ag_sheets = sheets_by_agniveer.get(agniveer.id, [])
            if not ag_sheets:
                continue
            
            sheet_map = {s.test_type: s for s in ag_sheets}
            sheet = sheet_map.get('CLK_FINAL')
            if not sheet:
                continue
                
            marks = _marks_from_sheet(sheet)
            tech_online = _num(marks.get('Tech Online (115)'))
            tech_proj = _num(marks.get('Tech Proj HRMS (25)'))
            academic = _num(marks.get('Academic Online (85)'))
            comp_online = _num(marks.get('Computer Online (25)'))
            comp_prac = _num(marks.get('Computer Prac (25)'))
            comp_total = _num(marks.get('Computer Total (50)')) or (comp_online + comp_prac)
            extempore = _num(marks.get('Extempore (25)'))
            typing_20 = marks.get('Typing 20 WPM', '')
            marks_obtained_300 = _num(marks.get('Marks Obtained (300)')) or (tech_online + tech_proj + academic + comp_total + extempore)
            percentage = round((marks_obtained_300 / 300) * 100, 2) if marks_obtained_300 else 0
            converted_40 = _num(marks.get('Marks Obtained (40)')) or round((marks_obtained_300 / 300) * 40, 2)
            is_pass = percentage >= 46
            
            if status_filter == 'evaluated' or is_pass == is_pass_filter:
                rows.append({
                    'army_no': agniveer.agniveer_no or agniveer.enrollment_number,
                    'rank': getattr(agniveer, 'rank', '') or '',
                    'trade': agniveer.trade or '',
                    'name': agniveer.get_full_name(),
                    'unit': agniveer.bn_desp or '',
                    'tech_online': tech_online,
                    'tech_proj': tech_proj,
                    'academic': academic,
                    'comp_online': comp_online,
                    'comp_prac': comp_prac,
                    'comp_total': comp_total,
                    'extempore': extempore,
                    'typing_20': typing_20,
                    'marks_obtained_300': marks_obtained_300,
                    'percentage': percentage,
                    'result_str': 'PASS' if is_pass else 'FAIL',
                    'grading': get_grade(percentage, [(tech_online, 115), (tech_proj, 25), (academic, 85), (comp_online, 25), (comp_prac, 25), (extempore, 25)], passing_pct=46),
                    'converted_40': converted_40,
                    'remarks': sheet.remarks if sheet.remarks else ''
                })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"CTS {status_filter.upper()}"
        st = _xl_styles(is_fail_theme=(status_filter == 'fail'))

        num_cols = 18 if status_filter == 'fail' else 19
        RESULT_COL = 16
        NAME_COL = 4

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        tc = ws.cell(row=1, column=1, value=f"FINAL RESULT OF CTS – {status_filter.upper()} LIST")
        tc.font = st['title_font']
        tc.fill = st['title_fill']
        tc.alignment = st['center']
        ws.row_dimensions[1].height = 34

        # Row 2 Category Headers
        row2_vals = {
            1: 'ARMY NO', 2: 'RANK', 3: 'TRADE', 4: 'NAME', 5: 'UNIT',
            6: 'Tech', 8: 'Academic', 9: 'CMPTR',
            12: 'Extempore (25/10)', 13: 'Typing 20 WPM',
            14: 'Marks Obtained (MM 300) 46%/138', 15: '%AGE', 16: 'RESULT',
        }
        if status_filter != 'fail':
            row2_vals.update({17: 'Grading', 18: 'Converted Out of 40', 19: 'Remarks'})
        else:
            row2_vals.update({17: 'Converted Out of 40', 18: 'Remarks'})

        # Row 3 Sub-headers
        row3_vals = {
            6: 'Online (115/46)', 7: 'Tech Proj (HRMS) (25/10)',
            8: 'Online (85/34)', 9: 'Online (25/10)',
            10: 'Prac (25/10)', 11: 'Total (50/20)',
        }

        for row_num, val_map in [(2, row2_vals), (3, row3_vals)]:
            ws.row_dimensions[row_num].height = 26
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = st['header_fill']
                cell.font = st['header_font']
                cell.alignment = st['center']
                cell.border = st['border']
                if col in val_map:
                    cell.value = val_map[col]

        # Vertical merges (row 2 and 3 for standalone columns)
        v_merges = [1, 2, 3, 4, 5, 12, 13, 14, 15, 16]
        if status_filter != 'fail':
            v_merges.extend([17, 18, 19])
        else:
            v_merges.extend([17, 18])
        for col in v_merges:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

        # Horizontal merges
        ws.merge_cells('F2:G2')   # Tech
        ws.merge_cells('I2:K2')   # CMPTR

        # Data rows start at row 4
        for index, r in enumerate(rows, 1):
            is_pass = r['result_str'] == 'PASS'
            values = [
                r['army_no'], r['rank'], r['trade'], r['name'], r['unit'],
                r['tech_online'], r['tech_proj'], r['academic'],
                r['comp_online'], r['comp_prac'], r['comp_total'],
                r['extempore'], r['typing_20'], r['marks_obtained_300'],
                f"{r['percentage']}%" if r['percentage'] else '0%', r['result_str'],
            ]
            if status_filter != 'fail':
                values.append(r['grading'])
            values.extend([r['converted_40'], r['remarks']])
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=index + 3, column=col)
                _xl_style_data_cell(cell, value, is_pass,
                                    is_result_col=(col == RESULT_COL),
                                    is_name_col=(col == NAME_COL), styles=st)
            ws.row_dimensions[index + 3].height = 20

        for col_idx in range(1, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            vals = [ws.cell(row=r, column=col_idx).value for r in range(2, len(rows) + 4)]
            ws.column_dimensions[col_letter].width = max((len(str(v or '')) for v in vals), default=10) + 4
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f"A3:{openpyxl.utils.get_column_letter(num_cols)}{max(len(rows)+3, 3)}"


        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_action(user, 'EXPORT', f'Exported Clerk {status_filter} dashboard results Excel', request)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="clerk_{status_filter}_results.xlsx"'
        return response

    def _export_specific_test_results(self, request, dept_code, sub_dept_key, test_type, status_filter):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from django.db.models import Q
        from django.http import HttpResponse
        from departments.models import Agniveer
        from evaluation.models import EvaluationSheet
        from evaluation.constants import DEPT_CONFIG
        from logs.utils import log_action

        CLERK_TRADES = ['CLK', 'CLERK', 'Clerk', 'CLK_SD', 'CLK_IM']

        # ── Resolve config ─────────────────────────────────────────────────────
        config = DEPT_CONFIG.get(dept_code, {})

        # For TTS (dept B) auto-detect sub_dept from test_type if not given
        effective_sub_dept = sub_dept_key
        if dept_code == 'B' and (not effective_sub_dept or effective_sub_dept == 'all'):
            if test_type.startswith('DMV_'):
                effective_sub_dept = 'DMV'
            elif test_type.startswith('OPEM_'):
                effective_sub_dept = 'OPEM'
            elif test_type.startswith('OTHER_'):
                effective_sub_dept = 'OTHER'

        # Resolve the right config block for sub_events / test label
        if dept_code == 'B' and effective_sub_dept in ['DMV', 'OPEM', 'OTHER']:
            tt_config = config.get('sub_departments', {}).get(effective_sub_dept, {})
        else:
            tt_config = config

        sub_events = tt_config.get('sub_events', {}).get(test_type, [])

        # Resolve human-readable test label
        test_type_label = test_type
        for tt_val, tt_lbl in tt_config.get('test_types', []):
            if tt_val == test_type:
                test_type_label = tt_lbl
                break
        if test_type_label == test_type:
            for tt_val, tt_lbl in config.get('test_types', []):
                if tt_val == test_type:
                    test_type_label = tt_lbl
                    break

        # ── Detect multi-attempt (Dept A tests with columns) ─────────────────
        # e.g. PPT/BPET/Firing/DST/BFC store results as {"1st Attempt": {…}, "Event Wise Best": {…}}
        dept_a_multi_attempt = (
            dept_code == 'A' and
            bool(config.get('test_config', {}).get(test_type, {}).get('columns'))
        )
        best_col_key = None
        if dept_a_multi_attempt:
            columns = config['test_config'][test_type].get('columns', [])
            best_col_key = next((c for c in columns if 'best' in c.lower()), columns[-1] if columns else None)

        # ── Agniveer scope ────────────────────────────────────────────────────
        all_agniveers = Agniveer.objects.all()
        if dept_code == 'A':
            if effective_sub_dept and effective_sub_dept != 'all':
                agniveers_qs = all_agniveers.filter(get_bn_desp_q('bn_desp', effective_sub_dept))
            else:
                agniveers_qs = all_agniveers
        elif dept_code == 'B':
            if effective_sub_dept == 'DMV':
                agniveers_qs = all_agniveers.filter(trade='DMV')
            elif effective_sub_dept == 'OPEM':
                agniveers_qs = all_agniveers.filter(trade='OPEM')
            elif effective_sub_dept == 'OTHER':
                agniveers_qs = all_agniveers.exclude(trade__in=['DMV', 'OPEM'] + CLERK_TRADES)
            else:
                agniveers_qs = all_agniveers.exclude(trade__in=CLERK_TRADES)
        elif dept_code == 'C':
            agniveers_qs = all_agniveers
        elif dept_code == 'D':
            agniveers_qs = all_agniveers.filter(trade__in=CLERK_TRADES)
        else:
            agniveers_qs = all_agniveers.none()

        sheets_qs = EvaluationSheet.objects.filter(
            department=dept_code,
            test_type=test_type,
            agniveer__in=agniveers_qs
        ).select_related('agniveer').prefetch_related('marks')

        # ── Filter rows by pass/fail/evaluated ────────────────────────────────
        filtered_rows = []
        for agniveer in agniveers_qs.order_by('agniveer_no', 'enrollment_number'):
            sheet = sheets_qs.filter(agniveer=agniveer).first()
            is_evaluated = sheet is not None
            is_pass = sheet.is_pass() if is_evaluated else False

            if status_filter == 'evaluated' and not is_evaluated:
                continue
            elif status_filter == 'pass' and (not is_evaluated or not is_pass):
                continue
            elif status_filter == 'fail' and (not is_evaluated or is_pass):
                continue

            filtered_rows.append({
                'agniveer': agniveer,
                'sheet': sheet,
                'is_evaluated': is_evaluated,
                'is_pass': is_pass
            })

        import re  # needed by _extract_event_marks

        dept_name = config.get('name', f'Dept {dept_code}')

        if dept_code == 'B' and effective_sub_dept in ['DMV', 'OPEM', 'OTHER']:
            dept_name = tt_config.get('name', effective_sub_dept)

        title_text = (
            f"ARMY EVALUATION PORTAL \u2014 {dept_name.upper()} \u2014 "
            f"{test_type_label.upper()} ({status_filter.upper()} LIST)"
        )

        # ── Build headers ────────────────────────────────────────────────────
        headers = ['S.No', 'Army No', 'Name', 'Trade', 'Unit']
        for event in sub_events:
            headers.append(event)
        headers.extend(['Total Marks', 'Result'])

        # ── Workbook setup ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = test_type[:30]
        st = _xl_styles(is_fail_theme=(status_filter == 'fail'))

        _xl_write_title(ws, title_text, len(headers), st)
        _xl_write_headers(ws, headers, st)

        RESULT_COL = len(headers)
        NAME_COL = 3

        unevaluated_fill = openpyxl.styles.PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

        def _extract_event_marks(sheet, sub_events, dept_a_multi_attempt, best_col_key):
            """Extract per-sub-event marks from the stored sub_event_results JSON."""
            res = sheet.sub_event_results or {}
            marks_dict = {}

            if dept_a_multi_attempt and best_col_key:
                # Multi-attempt format: {"1st Attempt": {event: val}, "Event Wise Best": {event: val}, ...}
                marks_dict = res.get(best_col_key, {})
                if not marks_dict:
                    # Fall back to any non-empty column
                    for col_data in res.values():
                        if isinstance(col_data, dict) and col_data:
                            marks_dict = col_data
                            break
            elif isinstance(res, dict):
                if isinstance(res.get('Marks'), dict):
                    marks_dict = res['Marks']
                else:
                    for ev in ['admin', 'officer', 'jco', 'nco']:
                        if isinstance(res.get(ev), dict) and res[ev]:
                            marks_dict = res[ev]
                            break
                    if not marks_dict:
                        marks_dict = res

            # Build a clean-key lookup (strip parenthetical annotations)
            clean_marks = {}
            if isinstance(marks_dict, dict):
                for k, v in marks_dict.items():
                    clean_k = re.sub(r'\s*\([^)]*\)', '', str(k)).strip()
                    clean_marks[clean_k] = v
                    clean_marks[k] = v  # keep original key too

            result = []
            for event in sub_events:
                val = clean_marks.get(event)
                if val is None:
                    event_clean = re.sub(r'\s*\([^)]*\)', '', event).strip().lower()
                    for ck, cv in clean_marks.items():
                        ck_clean = re.sub(r'\s*\([^)]*\)', '', str(ck)).strip().lower()
                        if ck_clean == event_clean or event_clean in ck_clean or ck_clean in event_clean:
                            val = cv
                            break
                result.append(val if val is not None else '\u2014')
            return result

        for row_idx, row_data in enumerate(filtered_rows, 1):
            agniveer = row_data['agniveer']
            sheet = row_data['sheet']
            is_eval = row_data['is_evaluated']
            is_pass = row_data['is_pass']

            values = [
                row_idx,
                agniveer.agniveer_no or agniveer.enrollment_number,
                agniveer.get_full_name(),
                agniveer.trade or '',
                agniveer.bn_desp or '',
            ]

            if is_eval:
                event_marks = _extract_event_marks(sheet, sub_events, dept_a_multi_attempt, best_col_key)
                values.extend(event_marks)
                values.extend([sheet.get_total_marks(), 'PASS' if is_pass else 'FAIL'])
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx + 2, column=col)
                    _xl_style_data_cell(cell, value, is_pass,
                                        is_result_col=(col == RESULT_COL),
                                        is_name_col=(col == NAME_COL), styles=st)
            else:
                for _ in sub_events:
                    values.append('')
                values.extend(['', 'NOT EVALUATED'])
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx + 2, column=col)
                    cell.value = value
                    cell.fill = unevaluated_fill
                    cell.alignment = st['left'] if col == NAME_COL else st['center']
                    cell.border = st['border']
                    cell.font = st['data_font']

            ws.row_dimensions[row_idx + 2].height = 20

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = (
            f"A2:{openpyxl.utils.get_column_letter(len(headers))}"
            f"{max(len(filtered_rows) + 2, 2)}"
        )
        for col_idx in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or ''))
                 for r in range(2, len(filtered_rows) + 3)),
                default=10
            )
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_action(request.user, 'EXPORT', f'Exported {test_type_label} Excel results ({status_filter})', request)

        filename = f"{dept_code}_{effective_sub_dept or 'all'}_{test_type}_{status_filter}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return HttpResponse("openpyxl not installed.", status=500)
        status_filter = request.GET.get('status', 'pass')
        if status_filter not in ['pass', 'fail', 'eligible', 'evaluated']:
            status_filter = 'pass'
        is_pass_filter = status_filter == 'pass'
        
        user = request.user

        dept_code = request.GET.get('dept')
        sub_dept_key = request.GET.get('sub_dept')
        test_type = request.GET.get('test_type')

        if dept_code and test_type:
            return self._export_specific_test_results(request, dept_code, sub_dept_key, test_type, status_filter)

        # Resolve target department (GET parameter takes precedence)
        target_dept = dept_code
        if not target_dept and user.is_department:
            target_dept = user.get_department_code()

        if target_dept == 'A':
            return self._export_battalion_results(request, status_filter, sub_dept=sub_dept_key)
        if target_dept == 'B':
            return self._export_tts_results(request, status_filter, sub_dept=sub_dept_key)
        if target_dept == 'C':
            return self._export_cs_results(request, status_filter, sub_dept=sub_dept_key)
        if target_dept == 'D':
            return self._export_clerk_results(request, status_filter, sub_dept=sub_dept_key)

        # Commander / G-Head with no dept specified → show all-battalion results (same as Battalion dashboard)
        return self._export_battalion_results(request, status_filter, sub_dept='all')


def extract_sheet_sub_events(sheet):
    if not sheet:
        return []
    res = sheet.sub_event_results or {}
    marks_dict = {}
    if isinstance(res, dict):
        if isinstance(res.get('Marks'), dict):
            marks_dict = res['Marks']
        else:
            for ev in ['admin', 'officer', 'jco', 'nco']:
                if isinstance(res.get(ev), dict) and res[ev]:
                    marks_dict = res[ev]
                    break
            if not marks_dict:
                marks_dict = res

    if not isinstance(marks_dict, dict):
        return []

    import re
    sub_events = []
    exclude_patterns = ['total', 'percentage', 'result', 'grading', 'status', 'remarks', 'convert', 'grand', 'round']

    for k, v in marks_dict.items():
        k_lower = k.lower()
        if any(p in k_lower for p in exclude_patterns):
            continue
        if k in ['Marks', 'admin', 'nco', 'jco', 'officer']:
            continue
            
        match = re.search(r'\((MM\s*|Max\s*Marks\s*)?(\d+)[^)]*\)', k)
        if match:
            max_val = int(match.group(2))
            clean_k = re.sub(r'\((MM\s*|Max\s*Marks\s*)?(\d+)[^)]*\)', '', k).strip()
        else:
            max_val = '—'
            clean_k = k.strip()

        try:
            val_float = float(v)
            val_str = f"{val_float:.1f}"
        except (ValueError, TypeError):
            val_str = str(v) if v is not None else '—'

        sub_events.append({
            'name': clean_k,
            'score': val_str,
            'max': max_val
        })
    return sub_events


class ExportPDFReportCardView(AnyStaffMixin, View):
    """Export individual Agniveer report card as PDF with enhanced ReportLab styling in Portrait layout."""

    def get(self, request, pk):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            return HttpResponse("reportlab not installed.", status=500)

        agniveer = get_object_or_404(Agniveer, pk=pk)
        evaluations = EvaluationSheet.objects.filter(agniveer=agniveer).prefetch_related('marks').order_by('department', 'category', 'test_type')
        all_evaluations = list(evaluations)

        def get_universal_dept_name_local(dept_code):
            if not dept_code:
                return ""
            dept_code = str(dept_code).strip().upper()
            trade = str(agniveer.trade or '').strip().upper()
            
            if dept_code == 'A':
                bn = str(agniveer.bn_desp or '').strip().lower()
                if '1tb' in bn:
                    return "1tb"
                elif '2tb' in bn or '2b' in bn:
                    return "2tb"
                return "1tb"
            elif dept_code == 'B':
                if trade == 'DMV':
                    return "dmv department"
                elif trade == 'OPEM':
                    return "opem"
                else:
                    return "tts department"
            elif dept_code == 'C':
                return "CES DEpartment"
            elif dept_code == 'D':
                return "cts department"
            return dept_code

        def get_final_sheet_for_dept(all_sheets, dept, trade):
            if dept == 'A':
                final_test_types = ['FINAL_RESULT']
            elif dept == 'B':
                if trade == 'DMV':
                    final_test_types = ['DMV_RESULT']
                elif trade == 'OPEM':
                    final_test_types = ['OPEM_RESULT']
                else:
                    final_test_types = ['OTHER_SCREEN_BOARD']
            elif dept == 'C':
                from evaluation.constants import CS_CLERK_RESULT_TRADES
                if trade in CS_CLERK_RESULT_TRADES:
                    final_test_types = ['CS_CLERK_RESULT']
                else:
                    final_test_types = ['CS_RESULT']
            elif dept == 'D':
                final_test_types = ['CLK_FINAL']
            else:
                final_test_types = []
                
            for sheet in all_sheets:
                if sheet.department == dept and sheet.test_type in final_test_types:
                    return sheet
            return None

        if request.user.is_trainer:
            return HttpResponse("You don't have permission to export report cards.", status=403)

        # Distinct departments in evaluations that actually have records
        available_depts = list(EvaluationSheet.objects.filter(agniveer=agniveer).values_list('department', flat=True).distinct())
        available_depts.sort()

        # Determine filtering based on role and optional query parameter
        target_dept = request.GET.get('dept')
        if not target_dept and (request.user.is_commander or request.user.is_g_head):
            if 'A' in available_depts:
                target_dept = 'A'
            elif available_depts:
                target_dept = available_depts[0]
            else:
                target_dept = 'A'

        is_department_export = False
        user_dept = None

        if request.user.is_department:
            user_dept = request.user.get_department_code()
            if not user_can_access_agniveer(request.user, agniveer, user_dept):
                return HttpResponse("You don't have permission to export this report card.", status=403)
            evaluations = scoped_sheets(evaluations, request.user, user_dept)
            is_department_export = True
        elif target_dept and target_dept in ['A', 'B', 'C', 'D'] and (request.user.is_commander or request.user.is_g_head):
            evaluations = evaluations.filter(department=target_dept)
            is_department_export = True
            user_dept = target_dept
        else:
            is_department_export = False
            user_dept = 'A'

        # Fetch active final sheets
        active_final_sheets = []
        if is_department_export:
            sheet = get_final_sheet_for_dept(all_evaluations, user_dept, agniveer.trade)
            if sheet:
                active_final_sheets.append(sheet)
        else:
            for d in ['A', 'B', 'C', 'D']:
                sheet = get_final_sheet_for_dept(all_evaluations, d, agniveer.trade)
                if sheet:
                    active_final_sheets.append(sheet)

        # Compute results
        from evaluation.result_helpers import build_department_result_row
        
        if is_department_export:
            department_result_row = build_department_result_row(agniveer, list(evaluations), user_dept)
            grand_total = department_result_row.get('grand_total', 0)
            max_total = department_result_row.get('max_total') or (120 if user_dept == 'A' else 40)
            percentage = department_result_row.get('percentage', 0)
            overall_pass = department_result_row.get('is_pass', False)
        else:
            department_result_row = {}
            # Compute summary for all 4 departments
            from evaluation.result_helpers import build_department_result_row, is_sheet_evaluated
            
            trade = str(agniveer.trade or '').strip().upper()
            bn_raw = str(agniveer.bn_desp or '').strip().lower()
            if '1tb' in bn_raw:
                bn_name = '1tb'
            elif '2tb' in bn_raw or '2b' in bn_raw:
                bn_name = '2tb'
            else:
                bn_name = '1tb'
                
            dept_name_map = {
                'A': bn_name,
                'B': 'dmv department' if trade == 'DMV' else ('opem' if trade == 'OPEM' else 'tts department'),
                'C': 'CES DEpartment',
                'D': 'cts department',
            }
            
            all_dept_results = {}
            for d in ['A', 'B', 'C', 'D']:
                d_sheets = [s for s in all_evaluations if s.department == d]
                if d_sheets:
                    d_row = build_department_result_row(agniveer, d_sheets, d)
                    is_eval = any(is_sheet_evaluated(s) for s in d_sheets)
                    if is_eval:
                        all_dept_results[d] = {
                            'name': dept_name_map.get(d, d),
                            'grand_total': d_row.get('grand_total', 0),
                            'max_total': d_row.get('max_total') or (120 if d == 'A' else 40),
                        }
                    else:
                        all_dept_results[d] = {
                            'name': dept_name_map.get(d, d),
                            'grand_total': '—',
                            'max_total': (120 if d == 'A' else 40),
                        }
                else:
                    all_dept_results[d] = {
                        'name': dept_name_map.get(d, d),
                        'grand_total': '—',
                        'max_total': (120 if d == 'A' else 40),
                    }
            
            grand_total = 0.0
            max_total = 0.0
            for d, res in all_dept_results.items():
                if res['grand_total'] != '—':
                    grand_total += float(res['grand_total'])
                    max_total += float(res['max_total'])
            percentage = round((grand_total / max_total * 100), 2) if max_total else 0
            overall_pass = percentage >= 50

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=0.4*inch, 
            bottomMargin=0.4*inch, 
            leftMargin=0.5*inch, 
            rightMargin=0.5*inch
        )
        styles = getSampleStyleSheet()

        # Brand colors matching the Web App
        brand_dark = colors.HexColor('#1B4332')
        brand_medium = colors.HexColor('#2D6A4F')
        gold_color = colors.HexColor('#D4A017')
        text_dark = colors.HexColor('#1A2C3E')
        text_muted = colors.HexColor('#5A6A7A')
        pass_green = colors.HexColor('#2E7D32')
        fail_red = colors.HexColor('#C62828')
        light_bg = colors.HexColor('#F7FAF8')
        border_color = colors.HexColor('#D1E2D6')

        elements = []

        # styles
        title_style = ParagraphStyle('Title', fontSize=14, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
        subtitle_style = ParagraphStyle('Sub', fontSize=7, textColor=colors.HexColor('#D4A017'), fontName='Helvetica', alignment=TA_CENTER, letterSpacing=2)
        badge_style = ParagraphStyle('Badge', fontSize=10, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
        badge_sub_style = ParagraphStyle('BadgeSub', fontSize=7, textColor=colors.white, fontName='Helvetica', alignment=TA_CENTER)
        
        # Profile styles
        label_style = ParagraphStyle('Label', fontSize=7, textColor=text_muted, fontName='Helvetica-Bold', leading=8)
        value_style = ParagraphStyle('Value', fontSize=9, textColor=text_dark, fontName='Helvetica-Bold', leading=11)

        # Table cell styles
        cell_style = ParagraphStyle('Cell', fontSize=7, fontName='Helvetica', alignment=TA_CENTER, leading=8)
        cell_bold_style = ParagraphStyle('CellBold', fontSize=7, fontName='Helvetica-Bold', alignment=TA_CENTER, leading=8)

        # ==== HEADER BANNER ====
        result_text = 'PASSED' if overall_pass else 'FAILED'
        result_color = pass_green if overall_pass else fail_red
        
        badge_content = [
          [Paragraph(result_text, ParagraphStyle('RT', parent=badge_style, textColor=result_color))],
          [Paragraph(f'{percentage:.2f}%', ParagraphStyle('PC', parent=badge_sub_style, textColor=colors.HexColor('#1A2C3E' if overall_pass else '#FFFFFF')))]
        ]
        badge_table = Table(badge_content, colWidths=[1.3*inch])
        badge_table.setStyle(TableStyle([
          ('BACKGROUND', (0, 0), (-1, -1), colors.Color(result_color.red, result_color.green, result_color.blue, alpha=0.15)),
          ('BORDER', (0, 0), (-1, -1), 1.5, result_color),
          ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
          ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
          ('TOPPADDING', (0, 0), (-1, -1), 3),
          ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        dept_names_map = {
            'A': get_universal_dept_name_local('A'),
            'B': get_universal_dept_name_local('B'),
            'C': get_universal_dept_name_local('C'),
            'D': get_universal_dept_name_local('D'),
        }
        title_banner_text = dept_names_map.get(user_dept, 'ARMY EVALUATION PORTAL').upper() if is_department_export else 'ARMY EVALUATION PORTAL'
        header_data = [
          [
            Paragraph(title_banner_text, title_style),
            badge_table
          ],
          [
            Paragraph(f'AGNIVEER PERFORMANCE ASSESSMENT REPORT — BATCH {agniveer.batch}', subtitle_style),
            ''
          ]
        ]
        
        header_table = Table(header_data, colWidths=[5.77*inch, 1.5*inch])
        header_table.setStyle(TableStyle([
          ('BACKGROUND', (0, 0), (-1, -1), brand_dark),
          ('SPAN', (1, 0), (1, 1)),
          ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
          ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
          ('TOPPADDING', (0, 0), (-1, -1), 6),
          ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
          ('RIGHTPADDING', (1, 0), (1, 1), 10),
        ]))
        # Construct hierarchical meta string
        dept_name_str = dept_names_map.get(user_dept, 'ALL DEPARTMENTS').upper()
        company_str = (agniveer.company or '—').upper()
        platoon_str = (agniveer.platoon or '—').upper()
        bn_str = (agniveer.bn_desp or '—').upper()
        
        if is_department_export:
            hierarchy_str = f"DEPARTMENT-{dept_name_str} — {company_str} — PLATOON-{platoon_str}"
        else:
            hierarchy_str = f"BATTALION-{bn_str} — ALL DEPARTMENTS — {company_str} — PLATOON-{platoon_str}"
            
        hierarchy_style = ParagraphStyle('HierarchyPDF', fontSize=7.5, textColor=colors.HexColor('#2D6A4F'), fontName='Helvetica-Bold', alignment=TA_CENTER)
        
        elements.append(header_table)
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(hierarchy_str, hierarchy_style))
        elements.append(Spacer(1, 6))

        # ==== PROFILE GRID ====
        try:
            if agniveer.photo:
                img = Image(agniveer.photo.path, 0.75*inch, 0.75*inch)
            else:
                img = Paragraph('PHOTO<br/>N/A', label_style)
        except:
            img = Paragraph('N/A', label_style)

        profile_grid = [
            [
                Paragraph('FULL NAME', label_style), 
                Paragraph('RANK', label_style), 
                Paragraph('TRADE', label_style), 
                Paragraph('BATTALION / UNIT', label_style)
            ],
            [
                Paragraph(agniveer.get_full_name().upper(), value_style), 
                Paragraph(department_result_row.get('rank', '—').upper() if is_department_export else getattr(agniveer, 'rank', '—').upper(), value_style), 
                Paragraph((agniveer.trade or 'Other').upper(), value_style), 
                Paragraph((agniveer.bn_desp or '—').upper(), value_style)
            ],
            [Spacer(1, 3), Spacer(1, 3), Spacer(1, 3), Spacer(1, 3)],
            [
                Paragraph('AGNIVEER NO (ARMY NO)', label_style), 
                Paragraph('ENROLLMENT NO', label_style), 
                Paragraph('COMPANY', label_style), 
                Paragraph('PLATOON', label_style)
            ],
            [
                Paragraph((agniveer.agniveer_no or '—').upper(), value_style), 
                Paragraph(agniveer.enrollment_number.upper(), value_style), 
                Paragraph(company_str, value_style), 
                Paragraph((agniveer.platoon or '—').upper(), value_style)
            ],
        ]
        
        grid_table = Table(profile_grid, colWidths=[1.54*inch, 1.54*inch, 1.54*inch, 1.55*inch])
        grid_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
        ]))

        main_profile_table = Table([[img, grid_table]], colWidths=[1.1*inch, 6.17*inch])
        main_profile_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('BACKGROUND', (0, 0), (-1, -1), light_bg),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(main_profile_table)
        elements.append(Spacer(1, 8))

        # ==== DEPARTMENT EVALUATION TABLE ====
        # We replace this with cards for active final sheets
        for sheet in active_final_sheets:
            dept_name_str = get_universal_dept_name_local(sheet.department).upper()
            title_text = f"<b>{dept_name_str} FINAL RESULT</b>"
            total_val = sheet.get_total_marks()
            max_val = sheet.get_max_marks()
            total_str = f"{int(total_val)}" if isinstance(total_val, (int, float)) and float(total_val).is_integer() else f"{total_val:.1f}"
            max_str = f"{int(max_val)}" if isinstance(max_val, (int, float)) and float(max_val).is_integer() else f"{max_val:.1f}"
            score_text = f"<b>{total_str} / {max_str}</b>"
            
            # Simple borderless table for final card header
            header_para_left = Paragraph(title_text, ParagraphStyle('CardTitle', fontSize=9, textColor=brand_dark, fontName='Helvetica-Bold'))
            header_para_right = Paragraph(score_text, ParagraphStyle('CardScore', fontSize=9, textColor=pass_green, fontName='Helvetica-Bold', alignment=TA_RIGHT))
            
            card_header_table = Table([[header_para_left, header_para_right]], colWidths=[5.27*inch, 2.0*inch])
            card_header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('LINEBELOW', (0, 0), (-1, -1), 1, border_color),
            ]))
            elements.append(card_header_table)
            elements.append(Spacer(1, 4))
            
            # Extract components
            sub_events = extract_sheet_sub_events(sheet)
            if sub_events:
                # Table headers
                sheet_headers = [
                    Paragraph('<b>TEST / SECTION COMPONENT</b>', ParagraphStyle('THL', fontSize=7.5, fontName='Helvetica-Bold', textColor=text_muted, alignment=TA_LEFT)),
                    Paragraph('<b>SCORE</b>', ParagraphStyle('THC', fontSize=7.5, fontName='Helvetica-Bold', textColor=text_muted, alignment=TA_CENTER)),
                    Paragraph('<b>MAX</b>', ParagraphStyle('THR', fontSize=7.5, fontName='Helvetica-Bold', textColor=text_muted, alignment=TA_CENTER))
                ]
                
                sheet_data = [sheet_headers]
                t_style = [
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]
                t_style.append(('LINEBELOW', (0, 0), (-1, 0), 1, border_color))

                for idx, event in enumerate(sub_events, 1):
                    try:
                        max_float = float(event['max'])
                        max_str = f"{int(max_float)}" if max_float.is_integer() else f"{max_float:.1f}"
                    except (ValueError, TypeError):
                        max_str = str(event['max'])
                        
                    is_total_conv = False
                    ev_name_lower = event['name'].lower()
                    if 'total' in ev_name_lower or 'converted' in ev_name_lower or 'round' in ev_name_lower:
                        is_total_conv = True
                    
                    if is_total_conv:
                        t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F0F9F4'))) # light green background
                        name_style = ParagraphStyle(f'EvNameTotal_{sheet.test_type}_{idx}', fontSize=7.5, fontName='Helvetica-Bold', textColor=colors.HexColor('#2E7D32'), alignment=TA_LEFT, leading=9)
                        score_style = ParagraphStyle(f'EvScoreTotal_{sheet.test_type}_{idx}', fontSize=8.5, fontName='Helvetica-Bold', textColor=colors.HexColor('#2E7D32'), alignment=TA_CENTER)
                        max_style = ParagraphStyle(f'EvMaxTotal_{sheet.test_type}_{idx}', fontSize=7.5, fontName='Helvetica-Bold', textColor=colors.HexColor('#2E7D32'), alignment=TA_CENTER)
                    else:
                        name_style = ParagraphStyle(f'EvName_{sheet.test_type}_{idx}', fontSize=7.5, fontName='Helvetica', alignment=TA_LEFT, leading=9)
                        score_style = ParagraphStyle(f'EvScore_{sheet.test_type}_{idx}', fontSize=8.5, fontName='Helvetica-Bold', alignment=TA_CENTER)
                        max_style = ParagraphStyle(f'EvMax_{sheet.test_type}_{idx}', fontSize=7.5, fontName='Helvetica', textColor=text_muted, alignment=TA_CENTER)
                        
                    sheet_data.append([
                        Paragraph(event['name'].upper(), name_style),
                        Paragraph(f"<b>{event['score']}</b>", score_style),
                        Paragraph(max_str, max_style)
                    ])

                for r in range(1, len(sheet_data) - 1):
                    t_style.append(('LINEBELOW', (0, r), (-1, r), 0.5, colors.HexColor('#E5ECE7')))
                
                sheet_table = Table(sheet_data, colWidths=[4.27*inch, 1.5*inch, 1.5*inch])
                sheet_table.setStyle(TableStyle(t_style))
                elements.append(sheet_table)
                
            # Remarks at the bottom of the card
            if sheet.remarks:
                elements.append(Spacer(1, 4))
                remarks_text = f"<i>Remarks: {sheet.remarks}</i>"
                elements.append(Paragraph(remarks_text, ParagraphStyle('CardRemarks', fontSize=8, fontName='Helvetica-Oblique', textColor=text_muted, leftIndent=12)))
            
            elements.append(Spacer(1, 15))

        # ==== CONSOLIDATED 4-DEPARTMENT OVERALL TABLE ====
        if request.user.is_commander or request.user.is_g_head:
            from evaluation.result_helpers import is_sheet_evaluated
            all_dept_results = {}
            for d in ['A', 'B', 'C', 'D']:
                d_sheets = [s for s in all_evaluations if s.department == d]
                if d_sheets:
                    d_row = build_department_result_row(agniveer, d_sheets, d)
                    is_eval = any(is_sheet_evaluated(s) for s in d_sheets)
                    if is_eval:
                        all_dept_results[d] = {
                            'name': get_universal_dept_name_local(d),
                            'grand_total': d_row.get('grand_total', 0),
                            'max_total': d_row.get('max_total') or (120 if d == 'A' else 40),
                            'percentage': d_row.get('percentage', 0),
                            'grading': d_row.get('grading', '—'),
                            'is_pass': d_row.get('is_pass', False),
                            'status': 'PASSED' if d_row.get('is_pass') else 'FAILED'
                        }
                    else:
                        all_dept_results[d] = {
                            'name': get_universal_dept_name_local(d),
                            'grand_total': '—',
                            'max_total': (120 if d == 'A' else 40),
                            'percentage': '—',
                            'grading': '—',
                            'status': 'PENDING'
                        }
                else:
                    all_dept_results[d] = {
                        'name': get_universal_dept_name_local(d),
                        'grand_total': '—',
                        'max_total': (120 if d == 'A' else 40),
                        'percentage': '—',
                        'grading': '—',
                        'status': 'PENDING'
                    }
            
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("OVERALL FOUR-DEPARTMENT PERFORMANCE SUMMARY", ParagraphStyle('ConsTitle', fontSize=8, textColor=brand_dark, fontName='Helvetica-Bold', spaceAfter=4)))
            
            cons_headers = [
                Paragraph('<b>Department / Assessment Phase</b>', cell_bold_style),
                Paragraph('<b>Marks Obtained</b>', cell_bold_style),
                Paragraph('<b>Max Marks</b>', cell_bold_style),
                Paragraph('<b>Percentage</b>', cell_bold_style),
                Paragraph('<b>Grading</b>', cell_bold_style),
                Paragraph('<b>Status</b>', cell_bold_style)
            ]
            
            cons_data = [cons_headers]
            for code, res in all_dept_results.items():
                pct_val = f"{res['percentage']:.2f}%" if res['percentage'] != '—' else '—'
                cons_data.append([
                    Paragraph(res['name'], ParagraphStyle('ConsName', fontSize=7, fontName='Helvetica-Bold', alignment=TA_LEFT)),
                    Paragraph(str(res['grand_total']), cell_style),
                    Paragraph(str(res['max_total']), cell_style),
                    Paragraph(pct_val, cell_style),
                    Paragraph(str(res['grading']), cell_style),
                    Paragraph(res['status'], ParagraphStyle('ConsStatus', fontSize=7, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=pass_green if res['status'] == 'PASSED' else (fail_red if res['status'] == 'FAILED' else text_muted)))
                ])
                
            cons_table = Table(cons_data, colWidths=[2.5*inch, 1.0*inch, 1.0*inch, 1.0*inch, 0.77*inch, 1.0*inch])
            cons_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), brand_dark),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, border_color),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FDF9')]),
            ]))
            elements.append(cons_table)

        # ==== SIGNATURES & FOOTER ====
        elements.append(Spacer(1, 25))
        
        sig_data = [
            [
                Paragraph('_______________________<br/><b>Training Officer</b>', ParagraphStyle('Sig1', fontSize=8, fontName='Helvetica', alignment=TA_CENTER)),
                Paragraph('<br/><b>[ SEAL ]</b>', ParagraphStyle('Sig2', fontSize=8, fontName='Helvetica', alignment=TA_CENTER)),
                Paragraph('_______________________<br/><b>Officer Commanding</b>', ParagraphStyle('Sig3', fontSize=8, fontName='Helvetica', alignment=TA_CENTER))
            ]
        ]
        sig_table = Table(sig_data, colWidths=[2.42*inch, 2.42*inch, 2.43*inch])
        sig_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(sig_table)
        
        elements.append(Spacer(1, 15))
        footer_style = ParagraphStyle('Footer', fontSize=7, alignment=TA_CENTER, textColor=text_muted)
        elements.append(Paragraph(f'Generated by Army Evaluation Portal | Date: {timezone.now().strftime("%d %b %Y, %H:%M")} | Trainee Enrollment: {agniveer.enrollment_number}', footer_style))

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="ReportCard_{agniveer.enrollment_number}.pdf"'
        
        log_action(request.user, 'EXPORT', f'Exported PDF Report Card for {agniveer.enrollment_number}', request)
        return response


class ExportDepartmentPDFView(AnyStaffMixin, View):
    """Export department summary as PDF with polished styling."""

    def get(self, request, dept):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            return HttpResponse("reportlab not installed.", status=500)

        if request.user.is_department and request.user.get_department_code() != dept:
            return HttpResponse("You don't have permission to export this department.", status=403)

        agniveers = scoped_agniveers(
            Agniveer.objects.filter(evaluations__department=dept).distinct(),
            request.user,
            dept,
        )
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.4*inch, bottomMargin=0.4*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
        styles = getSampleStyleSheet()
        
        # Brand Colors
        brand_dark = colors.HexColor('#1B4332')
        text_muted = colors.HexColor('#6C757D')

        elements = []

        # Header
        title_style = ParagraphStyle('Title', fontSize=20, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
        header_table = Table([[Paragraph(f"ARMY EVALUATION PORTAL - DEPARTMENT {dept} SUMMARY", title_style)]], colWidths=[10*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), brand_dark),
            ('PADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))

        hdr_style = ParagraphStyle('HdrStyle', fontSize=9, fontName='Helvetica-Bold', textColor=colors.white, alignment=1)
        cell_style = ParagraphStyle('CellStyle', fontSize=9, fontName='Helvetica', alignment=1)
        left_cell_style = ParagraphStyle('LeftCellStyle', fontSize=9, fontName='Helvetica', alignment=0)

        headers = [
            Paragraph('<b>S.No</b>', hdr_style),
            Paragraph('<b>Enrollment No</b>', hdr_style),
            Paragraph('<b>Name</b>', hdr_style),
            Paragraph('<b>Batch</b>', hdr_style),
            Paragraph('<b>Total Score</b>', hdr_style),
            Paragraph('<b>Pass/Fail Status</b>', hdr_style),
            Paragraph('<b>Remarks</b>', hdr_style),
        ]
        data = [headers]

        for i, a in enumerate(agniveers, 1):
            a_sheets = scoped_sheets(
                EvaluationSheet.objects.filter(agniveer=a, is_locked=True).prefetch_related('marks'),
                request.user,
                dept,
            )
            result_row = build_department_result_row(a, list(a_sheets), dept)
            raw_score = result_row.get('grand_total', 0)
            total_score = result_row.get('round_figure_120') if dept == 'A' else raw_score
            
            if dept == 'A':
                total_score_str = f"{total_score:g}"
            else:
                total_score_str = f"{total_score:.2f}"
                
            is_pass = result_row.get('is_pass', False)
            pass_status = 'Pass' if is_pass else 'Fail'
            
            remarks_list = [sheet.remarks.strip() for sheet in a_sheets if sheet.remarks and sheet.remarks.strip()]
            remarks_str = "; ".join(remarks_list) if remarks_list else ""

            data.append([
                Paragraph(str(i), cell_style),
                Paragraph(a.enrollment_number or '', cell_style),
                Paragraph(a.get_full_name() or '', left_cell_style),
                Paragraph(a.batch or '', cell_style),
                Paragraph(total_score_str, cell_style),
                Paragraph(pass_status or '', cell_style),
                Paragraph(remarks_str or '', left_cell_style),
            ])

        table = Table(data, colWidths=[0.5*inch, 1.5*inch, 2.3*inch, 1.0*inch, 1.2*inch, 1.3*inch, 2.75*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_dark),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(f'Generated by Army Evaluation Portal | Date: {timezone.now().strftime("%d %b %Y")}', ParagraphStyle('Footer', fontSize=8, alignment=TA_CENTER, textColor=text_muted)))

        doc.build(elements)
        buffer.seek(0)

        log_action(request.user, 'EXPORT', f'Exported Department {dept} PDF', request)
        return FileResponse(buffer, as_attachment=True, filename=f"dept_{dept}_summary.pdf")


# ── PowerPoint Dashboard Export Views ─────────────────────────────────────────

class ExportCommanderDashboardPPTXView(CommanderOrGHeadMixin, View):
    def get(self, request):
        import io
        from datetime import timedelta
        from django.utils import timezone
        from core.views import (
            build_evaluated_result_lists, 
            build_department_result_stats,
            DEPARTMENT_NAMES
        )
        from reports.pptx_generators import generate_commander_pptx

        # Key stats
        total_agniveers = Agniveer.objects.count()
        total_trainers = CustomUser.objects.filter(
            role__in=['trainer_nco', 'trainer_jco', 'trainer_officer']
        ).count()
        
        from evaluation.result_helpers import is_sheet_evaluated
        all_sheets = EvaluationSheet.objects.all().prefetch_related('marks')
        all_agniveers_qs = Agniveer.objects.prefetch_related('evaluations__marks')
        
        passed_agniveers, failed_agniveers = build_evaluated_result_lists(
            all_agniveers_qs,
            None,
            ['A'],
        )
        
        pass_count = len(passed_agniveers)
        fail_count = len(failed_agniveers)
        evaluated_agniveers = pass_count + fail_count
        evaluated_sheets_count = sum(1 for s in all_sheets if is_sheet_evaluated(s))
        completion_rate = (evaluated_sheets_count / max(total_agniveers * 28, 1)) * 100 if total_agniveers > 0 else 0
        pass_rate = (pass_count / max(evaluated_agniveers, 1)) * 100
        
        # Dept breakdown
        dept_labels = []
        dept_passed = []
        dept_failed = []
        for dept in ['A', 'B', 'C', 'D']:
            dept_sheets = all_sheets.filter(department=dept)
            dept_agniveers = Agniveer.objects.filter(evaluations__department=dept).distinct()
            stats = build_department_result_stats(dept, dept_sheets, dept_agniveers)
            dept_labels.append(DEPARTMENT_NAMES.get(dept, dept))
            dept_passed.append(stats['passed'])
            dept_failed.append(stats['failed'])
            
        # Monthly trend
        months = []
        month_counts = []
        month_pass = []
        for i in range(5, -1, -1):
            month_date = timezone.now() - timedelta(days=30 * i)
            month_sheets = all_sheets.filter(
                evaluation_date__year=month_date.year,
                evaluation_date__month=month_date.month
            )
            sheets_count = month_sheets.count()
            sheets_pass = sum(1 for s in month_sheets if s.is_pass())
            months.append(month_date.strftime('%b'))
            month_counts.append(sheets_count)
            month_pass.append(sheets_pass)
            
        # Top performers
        top_agniveers = []
        for agniveer in all_agniveers_qs[:20]:
            valid_sheets = [s for s in agniveer.evaluations.all() if is_sheet_evaluated(s)]
            if not valid_sheets:
                continue
            total_marks = sum(s.get_total_marks() for s in valid_sheets)
            max_marks = sum(s.get_max_marks() for s in valid_sheets)
            if max_marks > 0:
                percentage = (total_marks / max_marks) * 100
                top_agniveers.append({
                    'name': agniveer.get_full_name(),
                    'enrollment': agniveer.agniveer_no or agniveer.enrollment_number,
                    'percentage': round(percentage, 1)
                })
        top_agniveers = sorted(top_agniveers, key=lambda x: x['percentage'], reverse=True)

        data = {
            'total_agniveers': total_agniveers,
            'evaluated_agniveers': evaluated_agniveers,
            'pass_count': pass_count,
            'fail_count': fail_count,
            'pass_rate': round(pass_rate, 1),
            'completion_rate': round(completion_rate, 1),
            'dept_labels': dept_labels,
            'dept_passed': dept_passed,
            'dept_failed': dept_failed,
            'months_labels': months,
            'month_counts': month_counts,
            'month_pass': month_pass,
            'top_agniveers': top_agniveers,
        }
        
        prs = generate_commander_pptx(data)
        buffer = io.BytesIO()
        prs.save(buffer)
        buffer.seek(0)
        
        log_action(request.user, 'EXPORT', 'Exported Commander Dashboard PPTX', request)
        return FileResponse(
            buffer, 
            as_attachment=True, 
            filename="commander_dashboard_review.pptx",
            content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )


class ExportGHeadDashboardPPTXView(CommanderOrGHeadMixin, View):
    def get(self, request):
        import io
        from datetime import timedelta
        from django.utils import timezone
        from core.views import (
            build_evaluated_result_lists, 
            build_department_result_stats,
            DEPARTMENT_NAMES
        )
        from reports.pptx_generators import generate_ghead_pptx

        # Key stats
        total_agniveers = Agniveer.objects.count()
        total_trainers = CustomUser.objects.filter(
            role__in=['trainer_nco', 'trainer_jco', 'trainer_officer']
        ).count()
        
        from evaluation.result_helpers import is_sheet_evaluated
        all_sheets = EvaluationSheet.objects.all().prefetch_related('marks')
        all_agniveers_qs = Agniveer.objects.prefetch_related('evaluations__marks')
        
        passed_agniveers, failed_agniveers = build_evaluated_result_lists(
            all_agniveers_qs,
            None,
            ['A'],
        )
        
        pass_count = len(passed_agniveers)
        fail_count = len(failed_agniveers)
        evaluated_agniveers = pass_count + fail_count
        evaluated_sheets_count = sum(1 for s in all_sheets if is_sheet_evaluated(s))
        completion_rate = (evaluated_sheets_count / max(total_agniveers * 28, 1)) * 100 if total_agniveers > 0 else 0
        pass_rate = (pass_count / max(evaluated_agniveers, 1)) * 100
        
        # Dept breakdown
        dept_labels = []
        dept_passed = []
        dept_failed = []
        for dept in ['A', 'B', 'C', 'D']:
            dept_sheets = all_sheets.filter(department=dept)
            dept_agniveers = Agniveer.objects.filter(evaluations__department=dept).distinct()
            stats = build_department_result_stats(dept, dept_sheets, dept_agniveers)
            dept_labels.append(DEPARTMENT_NAMES.get(dept, dept))
            dept_passed.append(stats['passed'])
            dept_failed.append(stats['failed'])
            
        # Category breakdown
        category_labels = []
        category_pass_rates = []
        for category, label in EvaluationSheet.CATEGORY_CHOICES:
            cat_sheets = all_sheets.filter(category=category)
            if cat_sheets.exists():
                cat_pass = sum(1 for s in cat_sheets if s.is_pass())
                pr = (cat_pass / cat_sheets.count()) * 100
                category_labels.append(label)
                category_pass_rates.append(round(pr, 1))

        # Monthly trend
        months = []
        month_counts = []
        month_pass = []
        for i in range(5, -1, -1):
            month_date = timezone.now() - timedelta(days=30 * i)
            month_sheets = all_sheets.filter(
                evaluation_date__year=month_date.year,
                evaluation_date__month=month_date.month
            )
            sheets_count = month_sheets.count()
            sheets_pass = sum(1 for s in month_sheets if s.is_pass())
            months.append(month_date.strftime('%b'))
            month_counts.append(sheets_count)
            month_pass.append(sheets_pass)

        data = {
            'total_agniveers': total_agniveers,
            'evaluated_agniveers': evaluated_agniveers,
            'pass_count': pass_count,
            'fail_count': fail_count,
            'pass_rate': round(pass_rate, 1),
            'completion_rate': round(completion_rate, 1),
            'dept_labels': dept_labels,
            'dept_passed': dept_passed,
            'dept_failed': dept_failed,
            'category_labels': category_labels,
            'category_pass_rates': category_pass_rates,
            'months_labels': months,
            'month_counts': month_counts,
            'month_pass': month_pass,
        }
        
        prs = generate_ghead_pptx(data)
        buffer = io.BytesIO()
        prs.save(buffer)
        buffer.seek(0)
        
        log_action(request.user, 'EXPORT', 'Exported G-Head Dashboard PPTX', request)
        return FileResponse(
            buffer, 
            as_attachment=True, 
            filename="g_head_dashboard_review.pptx",
            content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )


class ExportDepartmentDashboardPPTXView(AnyStaffMixin, View):
    def get(self, request, dept):
        import io
        from datetime import timedelta
        from django.utils import timezone
        from core.views import (
            build_evaluated_result_lists, 
            get_scope_agniveers_and_sheets
        )
        from reports.pptx_generators import generate_department_pptx

        user = request.user
        if not (user.is_commander or user.is_g_head) and user.get_department_code() != dept:
            return HttpResponse("Unauthorized access to this department dashboard.", status=403)

        # Scoped data gathering
        all_sheets = EvaluationSheet.objects.all().prefetch_related('marks')
        all_agniveers = Agniveer.objects.all()
        
        sub_dept = 'all'
        if dept == 'A' and user.is_department and user.battalion_unit:
            sub_dept = user.battalion_unit
        elif dept == 'B' and user.is_department and user.tts_trade:
            sub_dept = user.tts_trade

        scoped_agniveers_qs, scoped_sheets_qs, total_eligible = get_scope_agniveers_and_sheets(
            dept, sub_dept, all_sheets.filter(department=dept), all_agniveers
        )

        from evaluation.result_helpers import is_sheet_evaluated
        passed_agniveers, failed_agniveers = build_evaluated_result_lists(
            scoped_agniveers_qs.prefetch_related('evaluations__marks'),
            None,
            [dept],
        )

        pass_count = len(passed_agniveers)
        fail_count = len(failed_agniveers)
        evaluated_agniveers = pass_count + fail_count
        
        evaluated_sheets_count = sum(1 for s in scoped_sheets_qs if is_sheet_evaluated(s))
        completion_rate = (evaluated_sheets_count / max(total_eligible * 28, 1)) * 100 if total_eligible > 0 else 0
        pass_rate = (pass_count / max(evaluated_agniveers, 1)) * 100

        # Category pass rates
        category_bar_labels = []
        category_pass_rates = []
        for category, label in EvaluationSheet.CATEGORY_CHOICES:
            cat_sheets = scoped_sheets_qs.filter(category=category)
            if cat_sheets.exists():
                cat_pass = sum(1 for s in cat_sheets if s.is_pass())
                category_bar_labels.append(label)
                category_pass_rates.append(round((cat_pass / cat_sheets.count()) * 100, 1))

        # Test type averages
        test_labels = []
        test_avg_marks = []
        test_pass_rates = []
        from evaluation.constants import get_dept_config
        user_config = get_dept_config(dept, user)
        for test_type, label in user_config.get('test_types', []):
            test_sheets = scoped_sheets_qs.filter(test_type=test_type)
            if test_sheets.exists():
                avg = sum(s.get_total_marks() for s in test_sheets) / test_sheets.count()
                pass_c = sum(1 for s in test_sheets if s.is_pass())
                pr = (pass_c / test_sheets.count()) * 100
            else:
                avg = 0.0
                pr = 0.0
            test_labels.append(label)
            test_avg_marks.append(round(avg, 1))
            test_pass_rates.append(round(pr, 1))

        # Monthly progress
        months = []
        month_counts = []
        month_pass = []
        for i in range(5, -1, -1):
            month_date = timezone.now() - timedelta(days=30 * i)
            month_sheets = scoped_sheets_qs.filter(
                evaluation_date__year=month_date.year,
                evaluation_date__month=month_date.month
            )
            sheets_count = month_sheets.count()
            sheets_pass = sum(1 for s in month_sheets if s.is_pass())
            months.append(month_date.strftime('%b'))
            month_counts.append(sheets_count)
            month_pass.append(sheets_pass)

        data = {
            'total_agniveers': total_eligible,
            'evaluated_agniveers': evaluated_agniveers,
            'pass_count': pass_count,
            'fail_count': fail_count,
            'pass_rate': round(pass_rate, 1),
            'completion_rate': round(completion_rate, 1),
            'category_bar_labels': category_bar_labels,
            'category_pass_rates': category_pass_rates,
            'test_labels': test_labels,
            'test_avg_marks': test_avg_marks,
            'test_pass_rates': test_pass_rates,
            'months_labels': months,
            'month_counts': month_counts,
            'month_pass': month_pass,
        }

        prs = generate_department_pptx(dept, data)
        buffer = io.BytesIO()
        prs.save(buffer)
        buffer.seek(0)

        log_action(request.user, 'EXPORT', f'Exported Department {dept} Dashboard PPTX', request)
        return FileResponse(
            buffer, 
            as_attachment=True, 
            filename=f"department_{dept}_dashboard_review.pptx",
            content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )


class ExportRegistrationDashboardPPTXView(LoginRequiredMixin, View):
    def get(self, request):
        import io
        from django.db.models import Count
        from departments.models import Agniveer, TRADE_CHOICES, COMPANY_CHOICES
        from reports.pptx_generators import generate_registration_pptx

        if not (request.user.is_commander or request.user.is_g_head or request.user.is_registration_office):
            return HttpResponse("Unauthorized", status=403)
            
        total_count = Agniveer.objects.count()
        
        # Trades breakdown
        trades_qs = Agniveer.objects.values('trade').annotate(count=Count('id'))
        trades_dict = {}
        for item in trades_qs:
            trade_code = item['trade']
            trade_label = trade_code
            for code, label in TRADE_CHOICES:
                if code == trade_code:
                    trade_label = label
                    break
            if not trade_label:
                trade_label = "Unassigned"
            trades_dict[trade_label] = item['count']
            
        # Company breakdown
        company_qs = Agniveer.objects.values('company').annotate(count=Count('id'))
        company_dict = {}
        for item in company_qs:
            c_code = item['company']
            c_label = c_code
            for code, label in COMPANY_CHOICES:
                if code == c_code:
                    c_label = label
                    break
            if not c_label:
                c_label = "Unassigned"
            company_dict[c_label] = item['count']
            
        recent_registrations = Agniveer.objects.order_by('-created_at')[:15]
        
        data = {
            'total_count': total_count,
            'trades_dict': trades_dict,
            'company_dict': company_dict,
            'recent_registrations': recent_registrations,
            'active_batch': 'Batch 2026',
        }
        
        prs = generate_registration_pptx(data)
        buffer = io.BytesIO()
        prs.save(buffer)
        buffer.seek(0)
        
        log_action(request.user, 'EXPORT', 'Exported Registration Dashboard PPTX', request)
        return FileResponse(
            buffer, 
            as_attachment=True, 
            filename="registration_dashboard_summary.pptx",
            content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )

