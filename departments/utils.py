"""
Departments App - Utilities
Excel parser for bulk Agniveer registration upload.
"""

import openpyxl
from datetime import date

# Maps Excel column header → model field name
EXCEL_COLUMN_MAP = {
    'Agniveer No': 'agniveer_no',
    'Name': 'name',
    'Father Name': 'father_name',
    'DOR': 'dor',
    'Trade': 'trade',
    'AROs/BROs': 'aros_bros',
    'BN DESP': 'bn_desp',
    'Batch No': 'batch_no',
    'Company': 'company',
    'Platoon': 'platoon',
    'Relationship': 'relationship',
    'AFMSF-2A': 'afmsf_2a',
    'Review Cert': 'review_cert',
    'EDN QL at time of enrolment': 'edn_ql_enrollment',
    'Higher EDN Qualification': 'higher_edn_qualification',
    'EDN Cert': 'edn_cert',
    'Verification Roll': 'verification_roll',
    'Character Cert': 'character_cert',
    'Unmarried Cert': 'unmarried_cert',
    'Caste Cert': 'caste_cert',
    'Class': 'class_field',
    'Domicile Cert': 'domicile_cert',
    'Outside Sanction Letter': 'outside_sanction_letter',
    'Willingness Cert': 'willingness_cert',
    'NCC Cert': 'ncc_cert',
    'Additional Cert': 'additional_cert',
    'PAN Card': 'pan_card',
    'Aadhar Card': 'aadhar_card',
    'Remarks': 'remarks',
}

YES_NO_FIELDS = {
    'afmsf_2a', 'review_cert', 'edn_cert', 'verification_roll',
    'character_cert', 'unmarried_cert', 'caste_cert', 'domicile_cert',
    'outside_sanction_letter', 'willingness_cert', 'ncc_cert',
    'pan_card', 'aadhar_card',
}


def get_excel_template_headers():
    """Return ordered list of headers for template download."""
    return list(EXCEL_COLUMN_MAP.keys())


def normalize_yes_no(value):
    """Normalise various truthy strings to Yes/No."""
    if value is None:
        return 'No'
    val = str(value).strip().lower()
    if val in ('yes', 'y', '1', 'true'):
        return 'Yes'
    return 'No'


def parse_agniveer_excel(file_obj):
    """
    Parse an uploaded Excel file and return (records, errors).

    records: list of dicts ready for model creation
    errors:  list of human-readable error strings
    """
    records = []
    errors = []

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return [], [f"Could not open Excel file: {e}"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["The Excel file is empty."]

    # Read header row (first row)
    headers = [str(h).strip() if h else '' for h in rows[0]]

    # Map column index → field name
    col_map = {}
    for idx, h in enumerate(headers):
        if h in EXCEL_COLUMN_MAP:
            col_map[idx] = EXCEL_COLUMN_MAP[h]

    if not col_map:
        return [], [
            "No recognised column headers found. "
            "Please download the template and use it exactly."
        ]

    for row_num, row in enumerate(rows[1:], start=2):
        data = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None

            if field_name in YES_NO_FIELDS:
                data[field_name] = normalize_yes_no(val)
            elif field_name == 'dor':
                if isinstance(val, (date,)):
                    data[field_name] = val
                elif val:
                    from datetime import datetime
                    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y'):
                        try:
                            data[field_name] = datetime.strptime(str(val).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        errors.append(f"Row {row_num}: Invalid date format for DOR: '{val}'")
                        data[field_name] = None
                else:
                    data[field_name] = None
            else:
                data[field_name] = str(val).strip() if val is not None else ''

        # Basic validation
        if not data.get('agniveer_no'):
            errors.append(f"Row {row_num}: 'Agniveer No' is required — skipped.")
            continue
        if not data.get('name'):
            errors.append(f"Row {row_num}: 'Name' is required — skipped.")
            continue

        records.append(data)

    wb.close()
    return records, errors


def generate_excel_template():
    """Generate an in-memory Excel workbook with correct headers for download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agniveer Registration"

    headers = get_excel_template_headers()
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(color="52B788", bold=True, size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)

    return wb
